import openpyxl

file_path = "PADRAO 8 HS.xlsx"
wb = openpyxl.load_workbook(file_path)

for sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
    for row in ws.iter_rows():
        for cell in row:
            if type(cell.value) is str and "DVEEDOR" in cell.value:
                cell.value = cell.value.replace("DVEEDOR", "DEVEDOR")

wb.save(file_path)
print("Fix applied to all sheets!")
