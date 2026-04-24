import openpyxl
import datetime

m_year = 2024
m_num = 1
d = 14

template_path = 'PADRAO 8 HS.xlsx'
wb = openpyxl.load_workbook(template_path)
ws = wb['JAN']

# Setting date
ws['K6'] = datetime.datetime(m_year, m_num, 1)

# Day 1 is Jan 1 -> Row 14
row_idx = 13 + 1
# Give it normal punches
ws.cell(row=row_idx, column=6, value=datetime.time(8, 0)).number_format = 'hh:mm:ss'
ws.cell(row=row_idx, column=7, value=datetime.time(12, 0)).number_format = 'hh:mm:ss'
ws.cell(row=row_idx, column=8, value=datetime.time(13, 0)).number_format = 'hh:mm:ss'
ws.cell(row=row_idx, column=9, value=datetime.time(17, 0)).number_format = 'hh:mm:ss'

# Day 2 is Jan 2 -> Row 15
row_idx = 15
# Give it an atestado
ws.cell(row=row_idx, column=12, value="ATESTADO")
ws.cell(row=row_idx, column=13, value="ATESTADO")
ws.cell(row=row_idx, column=14, value="ATESTADO")
ws.cell(row=row_idx, column=16, value=0)

# Day 3 is Jan 3 -> Row 16
row_idx = 16
# Give it a compensation
ws.cell(row=row_idx, column=6, value=datetime.time(8, 0)).number_format = 'hh:mm:ss'
ws.cell(row=row_idx, column=7, value=datetime.time(12, 0)).number_format = 'hh:mm:ss'
ws.cell(row=row_idx, column=8, value=datetime.time(13, 0)).number_format = 'hh:mm:ss'
ws.cell(row=row_idx, column=9, value=datetime.time(17, 0)).number_format = 'hh:mm:ss'
ws.cell(row=row_idx, column=10, value=datetime.time(17, 0)).number_format = 'hh:mm:ss'
ws.cell(row=row_idx, column=11, value=datetime.time(18, 0)).number_format = 'hh:mm:ss'

wb.save('TEST_OUTPUT_1.xlsx')
print("Saved TEST_OUTPUT_1.xlsx")
