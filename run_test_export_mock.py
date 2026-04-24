import os, datetime, calendar
import traceback
import sqlite3
import openpyxl

def rf(row, col):
    try:
        return row[col]
    except:
        return None

try:
    conn = sqlite3.connect('local.db')
    conn.row_factory = sqlite3.Row
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM TimeRecords LIMIT 10")
    rows = cursor.fetchall()
    
    if len(rows) == 0:
        print("No time records found to test.")
    else:
        target_mat = rows[0]['matricula']
        cursor.execute("SELECT * FROM TimeRecords WHERE matricula = ? ORDER BY timestamp DESC", (target_mat,))
        user_records = cursor.fetchall()
        user_records.reverse()
        
        workload_map = {target_mat: '40h'}
        cargo_map = {target_mat: 'Teste'}
        
        user_workload = workload_map.get(target_mat, '40h')
        daily_hours = 8
        
        m_year = datetime.datetime.now().year
        if user_records:
            ts = rf(user_records[-1], 'timestamp')
            dtt = ts if isinstance(ts, datetime.datetime) else datetime.datetime.strptime(str(ts).split('.')[0], '%Y-%m-%d %H:%M:%S')
            m_year = dtt.year
            
        months_data = {}
        for r in user_records:
            ts = rf(r, 'timestamp')
            dt = ts if isinstance(ts, datetime.datetime) else datetime.datetime.strptime(str(ts).split('.')[0], '%Y-%m-%d %H:%M:%S')
            if dt.year != m_year: continue
            month_key = dt.strftime('%Y-%m')
            if month_key not in months_data: months_data[month_key] = {"days": {}}
            day_key = dt.strftime('%Y-%m-%d')
            if day_key not in months_data[month_key]["days"]: months_data[month_key]["days"][day_key] = []
            months_data[month_key]["days"][day_key].append({'type': rf(r, 'record_type'), 'time': dt})

        template_path = "PADRAO 8 HS.xlsx"
        wb = openpyxl.load_workbook(template_path)
        
        month_names = ["JAN", "FEV", "MAR", "ABR", "MAIO", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
        for m_idx, m_name in enumerate(month_names):
            m_num = m_idx + 1
            if m_name in wb.sheetnames:
                ws = wb[m_name]
                ws['K6'] = datetime.datetime(m_year, m_num, 1)
                m_key = f"{m_year}-{m_num:02d}"
                month_data = months_data.get(m_key, {"days": {}})
                num_days = calendar.monthrange(m_year, m_num)[1]
                
                for d_idx in range(1, num_days + 1):
                    row_idx = 13 + d_idx
                    d_key_str = f"{m_year}-{m_num:02d}-{d_idx:02d}"
                    punches = month_data["days"].get(d_key_str, [])
                    
                    has_atestado = False
                    has_abono = False
                    has_comp = False
                    ent_m, sai_m, ent_t, sai_t, ent_x, sai_x = None, None, None, None, None, None
                    
                    if len(punches) > 0:
                        punches.sort(key=lambda x: x['time'])
                        # Just iterating
                        for p in punches:
                            t_str = p['type'].lower()
                            t_val = p['time'].time() if isinstance(p['time'], datetime.datetime) else None
                            if not t_val: continue
                            
                            if 'atestado' in t_str: has_atestado = True
                            elif 'abono' in t_str: has_abono = True
                            elif 'compensação' in t_str or 'compensacao' in t_str:
                                has_comp = True
                                if not ent_x: ent_x = t_val
                                elif not sai_x: sai_x = t_val
                            elif 'entrada' in t_str:
                                if ent_x and not sai_x: sai_x = t_val
                                elif not ent_m: ent_m = t_val
                            elif ('saída almoço' in t_str or 'saida almoco' in t_str) and not sai_m: sai_m = t_val
                            elif ('volta almoço' in t_str or 'volta almoco' in t_str) and not ent_t: ent_t = t_val
                            elif ('saída' in t_str and not 'almo' in t_str):
                                if not sai_m and not ent_t: sai_m = t_val
                                elif not sai_t: sai_t = t_val
                    
                    if has_atestado or has_abono:
                        val_str = "ATESTADO" if has_atestado else "ABONO"
                        ws.cell(row=row_idx, column=12, value=val_str)
                        ws.cell(row=row_idx, column=13, value=val_str)
                        ws.cell(row=row_idx, column=14, value=val_str)
                        ws.cell(row=row_idx, column=16, value=0)
                    else:
                        if ent_m: ws.cell(row=row_idx, column=6, value=ent_m).number_format = 'hh:mm:ss'
                        if sai_m: ws.cell(row=row_idx, column=7, value=sai_m).number_format = 'hh:mm:ss'
                        if ent_t: ws.cell(row=row_idx, column=8, value=ent_t).number_format = 'hh:mm:ss'
                        if sai_t: ws.cell(row=row_idx, column=9, value=sai_t).number_format = 'hh:mm:ss'
                        if ent_x: ws.cell(row=row_idx, column=10, value=ent_x).number_format = 'hh:mm:ss'
                        if sai_x: ws.cell(row=row_idx, column=11, value=sai_x).number_format = 'hh:mm:ss'
        
        print("Success!")
except Exception as e:
    print("Caught Exception:")
    traceback.print_exc()
