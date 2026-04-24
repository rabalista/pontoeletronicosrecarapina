import openpyxl
import datetime

m_year = 2024
m_num = 1
d = 14

template_path = 'PADRAO 8 HS.xlsx'
wb = openpyxl.load_workbook(template_path)
ws = wb['JAN']

# Setting date
ws['K6'] = datetime.datetime(m_year, m_num, 1)

# Just 1 missing punch
row_idx = 13 + 1
ws.cell(row=row_idx, column=6, value=datetime.time(8, 0)).number_format = 'hh:mm:ss'

wb.save('TEST_OUTPUT_2.xlsx')
print("Saved TEST_OUTPUT_2.xlsx")
