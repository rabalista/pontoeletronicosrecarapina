from flask import Flask, request, jsonify, send_from_directory, render_template, send_file
from flask_cors import CORS
import os
import bcrypt
import jwt
import datetime
import pytz
import sys
import time
import threading
from functools import wraps

import sqlite3
from dotenv import load_dotenv
import openpyxl
from io import BytesIO
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
import calendar

try:
    import pymssql
except ImportError:
    pymssql = None

load_dotenv()

app = Flask(__name__, static_folder='static', template_folder='templates')
_secret = os.getenv('SECRET_KEY')
if not _secret:
    _secret = os.getenv('DB_PASSWORD', 'ponto_sre_carapina')
app.config['SECRET_KEY'] = _secret

# Allow all origins, all methods, and specifically allow our bypass headers
try:
    CORS(app, resources={r"/api/*": {"origins": "*"}}, allow_headers=["*", "Authorization", "Content-Type", "Bypass-Tunnel-Reminder", "X-Tunnel-Skip-Proxy-Warning"], methods=["GET", "POST", "OPTIONS", "PUT", "DELETE"])
except:
    pass

# Database Configuration
server = os.getenv('DB_SERVER')
database = os.getenv('DB_NAME')
username = os.getenv('DB_USER')
password = os.getenv('DB_PASSWORD')
sqlite_path = os.getenv('SQLITE_PATH', 'local.db')
USE_SQLITE = os.getenv('USE_SQLITE', 'true').lower() == 'true'

# Optional: initialize database/tables on startup when deploying
if os.getenv('INIT_DB_ON_START', 'false').lower() == 'true':
    try:
        from setup_db import create_database, create_tables
        create_database()
        create_tables()
    except Exception:
        pass

def ensure_sql_server_tables():
    # Only run if not using SQLite exclusively
    if not USE_SQLITE:
        try:
            import pyodbc
            conn_str = f'DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={db_server};DATABASE={db_name};UID={db_user};PWD={db_password};TrustServerCertificate=yes;Connection Timeout=5'
            conn = pyodbc.connect(conn_str)
            conn.autocommit = True
            cursor = conn.cursor()
            cursor.execute("""
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='CustomHolidays' and xtype='U')
            BEGIN
                CREATE TABLE CustomHolidays (
                    date_str NVARCHAR(10) PRIMARY KEY,
                    description NVARCHAR(255) NOT NULL
                )
            END
            IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='SystemConfig' and xtype='U')
            BEGIN
                CREATE TABLE SystemConfig (
                    [key] NVARCHAR(50) PRIMARY KEY,
                    [value] NVARCHAR(MAX)
                )
                INSERT INTO SystemConfig ([key], [value]) VALUES ('excel_protection_password', 'Sedu@2023')
            END
            """)
            conn.close()
        except Exception as e:
            try:
                import pymssql
                conn = pymssql.connect(server=db_server, user=db_user, password=db_password, database=db_name, timeout=5, autocommit=True)
                cursor = conn.cursor()
                cursor.execute("""
                IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='CustomHolidays' and xtype='U')
                BEGIN
                    CREATE TABLE CustomHolidays (
                        date_str NVARCHAR(10) PRIMARY KEY,
                        description NVARCHAR(255) NOT NULL
                    )
                END
                IF NOT EXISTS (SELECT * FROM sysobjects WHERE name='SystemConfig' and xtype='U')
                BEGIN
                    CREATE TABLE SystemConfig (
                        [key] NVARCHAR(50) PRIMARY KEY,
                        [value] NVARCHAR(MAX)
                    )
                    INSERT INTO SystemConfig ([key], [value]) VALUES ('excel_protection_password', 'Sedu@2023')
                END
                """)
                conn.close()
            except Exception as ex:
                print("Failed to ensure SQL Server tables:", ex)

# Ensure SQL Server tables on boot
threading.Thread(target=ensure_sql_server_tables).start()

# Global DB Status
DB_ONLINE = False

# Configuração SQL Server
db_server = os.getenv('DB_SERVER')
db_name = os.getenv('DB_NAME')
db_user = os.getenv('DB_USER')
db_password = os.getenv('DB_PASSWORD')

# Global Sync Locking
SYNC_LOCKS = {}
SYNC_LOCKS_LOCK = threading.Lock()

def get_user_sync_lock(matricula):
    with SYNC_LOCKS_LOCK:
        if matricula not in SYNC_LOCKS:
            SYNC_LOCKS[matricula] = threading.Lock()
        return SYNC_LOCKS[matricula]

def get_fifth_business_day(year, month):
    import datetime
    holidays = [(1, 1), (4, 21), (5, 1), (9, 7), (10, 12), (11, 2), (11, 15), (12, 25)]
    count = 0
    d = datetime.date(year, month, 1)
    while count < 5:
        if d.weekday() < 5 and (d.month, d.day) not in holidays:
            count += 1
        if count < 5:
            d += datetime.timedelta(days=1)
    return d

def get_remote_db_connection():
    """Explicitly tries to connect to SQL Server, ignoring USE_SQLITE."""
    try:
        import pyodbc
        conn_str = f'DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={db_server};DATABASE={db_name};UID={db_user};PWD={db_password};TrustServerCertificate=yes;Connection Timeout=5'
        conn = pyodbc.connect(conn_str)
        conn.autocommit = True
        return conn
    except Exception:
        try:
            import pymssql
            conn = pymssql.connect(server=db_server, user=db_user, password=db_password, database=db_name, timeout=5, as_dict=True, autocommit=True)
            return conn
        except Exception:
            return None

def get_db_connection():
    if USE_SQLITE:
        conn = sqlite3.connect(sqlite_path)
        conn.row_factory = sqlite3.Row
        ensure_sqlite_schema(conn)
        return conn
    
    # Try SQL Server
    try:
        import pyodbc
        conn_str = f'DRIVER={{ODBC Driver 18 for SQL Server}};SERVER={db_server};DATABASE={db_name};UID={db_user};PWD={db_password};TrustServerCertificate=yes;Connection Timeout=5'
        conn = pyodbc.connect(conn_str)
        conn.autocommit = True
        return conn
    except Exception:
        try:
            import pymssql
            conn = pymssql.connect(server=db_server, user=db_user, password=db_password, database=db_name, timeout=5, as_dict=True, autocommit=True)
            return conn
        except Exception:
            # Fallback to SQLite
            conn = sqlite3.connect(sqlite_path)
            conn.row_factory = sqlite3.Row
            ensure_sqlite_schema(conn)
            return conn

def check_db_status():
    global DB_ONLINE
    # Trigger an initial sync to clear any pending records after migration
    threading.Thread(target=auto_sync_all, daemon=True).start()
    
    while True:
        # Reverting to real SQL Server connection check
        if USE_SQLITE:
            DB_ONLINE = True
            time.sleep(60)
            continue

        try:
            conn = get_db_connection()
            # If it's a real SQL Server connection (not the sqlite fallback)
            if not isinstance(conn, sqlite3.Connection):
                DB_ONLINE = True
                conn.close()
            else:
                DB_ONLINE = False
                # If it fell back to sqlite, we are "OFFLINE" relative to SQL Server
        except Exception:
            DB_ONLINE = False
            
        time.sleep(5) # Probing every 5 seconds for "instant" feeling

def start_health_check():
    t = threading.Thread(target=check_db_status)
    t.daemon = True
    t.start()
    
def ensure_sqlite_schema(conn):
    c = conn.cursor()
    c.execute("""
        CREATE TABLE IF NOT EXISTS Users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            matricula TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            name TEXT NOT NULL,
            cargo TEXT DEFAULT 'Funcionario',
            role TEXT DEFAULT 'user',
            must_clear_cache INTEGER DEFAULT 0
        )
    """)
    # Migration for existing SQLite Users table
    try: c.execute("ALTER TABLE Users ADD COLUMN must_clear_cache INTEGER DEFAULT 0")
    except: pass
    try: c.execute("ALTER TABLE Users ADD COLUMN cargo TEXT DEFAULT 'Funcionario'")
    except: pass
    try: c.execute("ALTER TABLE Users ADD COLUMN workload TEXT DEFAULT '40h'")
    except: pass
    c.execute("""
        CREATE TABLE IF NOT EXISTS TimeRecords (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id INTEGER,
            matricula TEXT NOT NULL,
            record_type TEXT NOT NULL,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            neighborhood TEXT,
            city TEXT,
            latitude FLOAT,
            longitude FLOAT,
            accuracy FLOAT,
            full_address TEXT,
            user_name TEXT,
            transaction_id TEXT UNIQUE,
            cargo TEXT DEFAULT 'Funcionario'
        );
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS OfflineQueue (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            matricula TEXT NOT NULL,
            record_type TEXT NOT NULL,
            timestamp DATETIME DEFAULT CURRENT_TIMESTAMP,
            neighborhood TEXT,
            city TEXT,
            latitude FLOAT,
            longitude FLOAT,
            accuracy FLOAT,
            full_address TEXT,
            transaction_id TEXT UNIQUE,
            cargo TEXT DEFAULT 'Funcionario'
        );
    """)
    # Add columns if they don't exist
    new_cols = [
        ("TimeRecords", "latitude", "REAL"),
        ("TimeRecords", "longitude", "REAL"),
        ("TimeRecords", "accuracy", "REAL"),
        ("TimeRecords", "full_address", "TEXT"),
        ("OfflineQueue", "latitude", "REAL"),
        ("OfflineQueue", "longitude", "REAL"),
        ("OfflineQueue", "accuracy", "REAL"),
        ("OfflineQueue", "full_address", "TEXT"),
        ("TimeRecords", "matricula", "TEXT"),
        ("TimeRecords", "user_name", "TEXT"),
        ("OfflineQueue", "matricula", "TEXT"),
        ("OfflineQueue", "user_name", "TEXT"),
        ("TimeRecords", "transaction_id", "TEXT"),
        ("OfflineQueue", "transaction_id", "TEXT"),
        ("TimeRecords", "is_retroactive", "INTEGER"),
        ("TimeRecords", "justification", "TEXT"),
        ("TimeRecords", "document_path", "TEXT"),
        ("OfflineQueue", "is_retroactive", "INTEGER"),
        ("OfflineQueue", "justification", "TEXT"),
        ("OfflineQueue", "document_path", "TEXT"),
        ("TimeRecords", "is_reviewed", "INTEGER"),
        ("OfflineQueue", "is_reviewed", "INTEGER")
    ]
    for table, col, col_type in new_cols:
        try:
            c.execute(f"ALTER TABLE {table} ADD COLUMN {col} {col_type}")
        except: pass
        
    c.execute("""
        CREATE TABLE IF NOT EXISTS SystemConfig (
            key TEXT PRIMARY KEY,
            value TEXT
        )
    """)
    c.execute("""
        CREATE TABLE IF NOT EXISTS CustomHolidays (
            date_str TEXT PRIMARY KEY,
            description TEXT NOT NULL
        )
    """)
    # Default passwords
    c.execute("INSERT OR IGNORE INTO SystemConfig (key, value) VALUES ('location_edit_password', 'admin123')")
    c.execute("INSERT OR IGNORE INTO SystemConfig (key, value) VALUES ('excel_protection_password', 'Sedu@2023')")
    
    conn.commit()

def migrate_local_data():
    """Populates matricula and user_name in existing TimeRecords and OfflineQueue entries."""
    print("DEBUG: Starting local data migration...")
    try:
        conn = sqlite3.connect(sqlite_path)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        
        # Build mapping
        cur.execute("SELECT id, matricula, name FROM Users")
        user_map = {r['id']: (r['matricula'], r['name']) for r in cur.fetchall()}
        
        # Update TimeRecords
        cur.execute("SELECT id, user_id FROM TimeRecords WHERE matricula IS NULL")
        to_update = cur.fetchall()
        for r in to_update:
            if r['user_id'] in user_map:
                m, n = user_map[r['user_id']]
                cur.execute("UPDATE TimeRecords SET matricula = ?, user_name = ? WHERE id = ?", (m, n, r['id']))
        
        # Update OfflineQueue
        cur.execute("SELECT id, user_id FROM OfflineQueue WHERE matricula IS NULL")
        to_update_q = cur.fetchall()
        for r in to_update_q:
            if r['user_id'] in user_map:
                m, n = user_map[r['user_id']]
                cur.execute("UPDATE OfflineQueue SET matricula = ?, user_name = ? WHERE id = ?", (m, n, r['id']))
        
        conn.commit()
        conn.close()
        print("DEBUG: Local data migration complete.")
    except Exception as e:
        print(f"DEBUG: Migration error: {e}")
    # The original instruction had conn.commit() here, but it should be inside the try block before conn.close()
    # or removed if the previous commit covers it. Given the structure, it's likely a copy-paste error.
    # I'll remove the redundant conn.commit() here as it's already done inside the try block.

def ensure_default_admin():
    try:
        admin_mat = os.getenv('ADMIN_MATRICULA', 'admin')
        admin_pass = os.getenv('ADMIN_PASSWORD', 'admin')
        admin_name = os.getenv('ADMIN_NAME', 'Administrador')
        sconn = sqlite3.connect(sqlite_path)
        sconn.row_factory = sqlite3.Row
        ensure_sqlite_schema(sconn)
        scur = sconn.cursor()
        scur.execute("SELECT 1 FROM Users WHERE matricula = ?", (admin_mat,))
        exists = scur.fetchone()
        if not exists:
            hashed = bcrypt.hashpw(admin_pass.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
            try:
                scur.execute("INSERT OR IGNORE INTO Users (matricula, password, name, role) VALUES (?, ?, ?, ?)", (admin_mat, hashed, admin_name, 'admin'))
                sconn.commit()
            except Exception:
                pass
        sconn.close()
    except Exception:
        pass




def sql_online():
    return DB_ONLINE

def get_ph(conn):
    """Returns the correct SQL placeholder based on the connection type."""
    if isinstance(conn, sqlite3.Connection):
        return '?'
    if 'pymssql' in str(type(conn)):
        return '%s'
    return '?'

def get_user_info_by_id(user_id, conn):
    """Returns (matricula, name) for a given user_id using the provided connection."""
    ph = get_ph(conn)
    is_sqlite = isinstance(conn, sqlite3.Connection)
    try:
        cur = conn.cursor()
        nolock = "" if is_sqlite else "WITH (NOLOCK)"
        cur.execute(f"SELECT matricula, name FROM Users {nolock} WHERE id = {ph}", (user_id,))
        row = cur.fetchone()
        if row:
            return rf(row, 'matricula'), rf(row, 'name')
    except:
        pass
    return None, None

def get_user_info_by_matricula(matricula, conn):
    """Returns (id, name) for a given matricula using the provided connection."""
    ph = get_ph(conn)
    is_sqlite = isinstance(conn, sqlite3.Connection)
    try:
        cur = conn.cursor()
        nolock = "" if is_sqlite else "WITH (NOLOCK)"
        cur.execute(f"SELECT id, name FROM Users {nolock} WHERE matricula = {ph}", (matricula,))
        row = cur.fetchone()
        if row:
            return rf(row, 'id'), rf(row, 'name')
    except:
        pass
    return None, None


def rf(row, name):
    """
    Robustly fetch a column from a database row (sqlite3.Row, pyodbc.Row, dict, or tuple).
    """
    if row is None:
        return None
    try:
        # 1. Try dict-style access (sqlite3.Row, dict, pymssql as_dict)
        if hasattr(row, 'get'):
            return row.get(name)
        
        # 2. Try index/key access
        try:
            return row[name]
        except (TypeError, IndexError, KeyError):
            pass
            
        # 3. Try attribute access (pyodbc.Row)
        return getattr(row, name)
    except Exception:
        # 4. Final attempt: fallback to column indices if we know the row is a tuple and which column it is
        # This is a bit risky but helped if col order is fixed
        return None

# Auth Decorator
def token_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        auth = request.headers.get("Authorization", "") or ""
        parts = auth.split()
        token = parts[1] if len(parts) == 2 and parts[0].lower() == "bearer" else None
        if not token:
            return jsonify({"message": "Token is missing!"}), 401
        try:
            data = jwt.decode(token, app.config["SECRET_KEY"], algorithms=["HS256"])
            # Favor matricula for cross-system stability
            curr_user_mat = data.get("matricula")
            if not curr_user_mat:
                # Fallback for old tokens if any
                curr_user_mat = str(data.get("user_id"))
            role = data["role"]
        except Exception as e:
            print(f"❌ Token validation failed: {str(e)}")
            return jsonify({"message": "Token is invalid!"}), 401
        return f(curr_user_mat, role, *args, **kwargs)
    return decorated

# Routes
@app.route('/')
def index():
    return render_template('index.html')

@app.route('/index.html')
def index_html():
    return render_template('index.html')

@app.route('/config.js')
def serve_config():
    return send_from_directory('static', 'config.js')

@app.route('/register.html')
def register_html():
    return render_template('register.html')

@app.route('/dashboard.html')
def dashboard_html():
    return render_template('dashboard.html')

@app.route('/admin.html')
def admin_html():
    return render_template('admin.html')

@app.route('/health')
def health():
    return jsonify({'status': 'ok'}), 200

@app.route('/register')
def register_page():
    return render_template('register.html')

@app.route('/dashboard')
def dashboard():
    return render_template('dashboard.html')

@app.route('/admin')
def admin_page():
    return render_template('admin.html')

# API Endpoints
@app.route('/api/register', methods=['POST'])
def register():
    data = request.get_json()
    hashed_password = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    
    
    
    cargo = data.get('cargo')
    if not cargo:
        cargo = 'Funcionario'
    workload = data.get('workload', '40h')
    
    conn = get_db_connection()
    ph = get_ph(conn)
    cursor = conn.cursor()
    
    try:
        # Ensure 'cargo' and 'workload' column exists on SQL Server if using it
        if not isinstance(conn, sqlite3.Connection):
            try: 
                cursor.execute("""
                IF COL_LENGTH('Users', 'cargo') IS NULL BEGIN ALTER TABLE Users ADD cargo NVARCHAR(100) DEFAULT 'Funcionario' END
                """)
                cursor.execute("""
                IF COL_LENGTH('Users', 'workload') IS NULL BEGIN ALTER TABLE Users ADD workload NVARCHAR(50) DEFAULT '40h' END
                """)
                conn.commit()
            except: pass

        query = f"INSERT INTO Users (matricula, password, name, role, cargo, workload) VALUES ({ph}, {ph}, {ph}, {ph}, {ph}, {ph})"
        cursor.execute(query, (data['matricula'], hashed_password, data['name'], 'user', cargo, workload))
        try: conn.commit()
        except: pass
        # mirror to local sqlite for offline login
        try:
            sconn = sqlite3.connect(sqlite_path)
            sconn.row_factory = sqlite3.Row
            ensure_sqlite_schema(sconn)
            scur = sconn.cursor()
            scur.execute("INSERT OR IGNORE INTO Users (matricula, password, name, role, cargo, workload) VALUES (?, ?, ?, ?, ?, ?)",
                         (data['matricula'], hashed_password, data['name'], 'user', cargo, workload))
            sconn.commit()
            sconn.close()
        except Exception:
            pass
        return jsonify({'message': 'User registered successfully!'}), 201
    except Exception as e:
        msg = str(e)
        if 'UNIQUE' in msg or 'unique' in msg or 'duplicate' in msg:
            return jsonify({'message': 'Matricula already exists!'}), 409
        return jsonify({'message': msg}), 500
    finally:
        conn.close()

@app.route('/api/login', methods=['POST'])
def login():
    data = request.get_json()
    conn = get_db_connection()
    ph = get_ph(conn)
    user = None
    is_sqlite = isinstance(conn, sqlite3.Connection)
    try:
        cursor = conn.cursor()
        if not is_sqlite:
            query = f"SELECT id, matricula, password, name, role, cargo FROM Users WITH (NOLOCK) WHERE matricula = {ph}"
        else:
            query = f"SELECT id, matricula, password, name, role, cargo FROM Users WHERE matricula = {ph}"
        cursor.execute(query, (data['matricula'],))
        user = cursor.fetchone()
    except Exception:
        user = None
    finally:
        try:
            conn.close()
        except Exception:
            pass
    
    # Fallback to local sqlite if not found on primary connection (if primary was SQL)
    if not user and not is_sqlite:
        try:
            sconn = sqlite3.connect(sqlite_path)
            sconn.row_factory = sqlite3.Row
            ensure_sqlite_schema(sconn)
            scur = sconn.cursor()
            scur = sconn.cursor()
            scur.execute("SELECT id, matricula, password, name, role, cargo FROM Users WHERE matricula = ?", (data['matricula'],))
            user = scur.fetchone()
            sconn.close()
        except Exception:
            user = None
    
    pw_field = rf(user, 'password')
    if user and pw_field and bcrypt.checkpw(data['password'].encode('utf-8'), pw_field.encode('utf-8')):
        # Mirror user to local sqlite for future offline login
        try:
            sconn = sqlite3.connect(sqlite_path)
            sconn.row_factory = sqlite3.Row
            ensure_sqlite_schema(sconn)
            scur = sconn.cursor()
            current_hash = rf(user, 'password')
            
            scur.execute("SELECT 1 FROM Users WHERE matricula = ?", (data['matricula'],))
            exists = scur.fetchone()
            if exists:
                 scur.execute("UPDATE Users SET password = ?, name = ?, role = ?, cargo = ? WHERE matricula = ?", 
                              (current_hash, rf(user, 'name'), rf(user, 'role'), rf(user, 'cargo'), data['matricula']))
            else:
                scur.execute("INSERT INTO Users (matricula, password, name, role, cargo) VALUES (?, ?, ?, ?, ?)",
                             (data['matricula'], current_hash, rf(user, 'name'), rf(user, 'role'), rf(user, 'cargo')))
            sconn.commit()
            sconn.close()
        except Exception:
            pass

        try:
            token = jwt.encode({
                'matricula': rf(user, 'matricula'),
                'user_id': rf(user, 'id'), # keep for backward compat if needed
                'role': rf(user, 'role'),
                'exp': datetime.datetime.now(datetime.timezone.utc) + datetime.timedelta(hours=24)
            }, app.config['SECRET_KEY'], algorithm="HS256")
            
            # Trigger background sync for this user immediately
            threading.Thread(target=perform_sync_for_user, args=(rf(user, 'matricula'),), daemon=True).start()
            
            return jsonify({'token': token, 'role': rf(user, 'role'), 'name': rf(user, 'name'), 'cargo': rf(user, 'cargo')})
        except Exception as e:
            return jsonify({'message': f'Internal Server Error: {str(e)}'}), 500
    
    return jsonify({'message': 'Invalid credentials!'}), 401

@app.route('/api/punch', methods=['POST'])
@token_required
def punch(curr_user_mat, role):
    data = request.get_json()
    # Expecting: type, neighborhood, city
    
    conn = get_db_connection()
    # Determine basic status
    is_sqlite = isinstance(conn, sqlite3.Connection)
    ph = get_ph(conn)

    current_time_raw = datetime.datetime.now(pytz.timezone('America/Sao_Paulo')).replace(tzinfo=None)
    current_time = current_time_raw.replace(microsecond=0) # Standardize to seconds
    transaction_id = data.get('transaction_id')
    
    # fetch denormalized user fields if available
    user_matricula = curr_user_mat
    sql_user_id, user_name = get_user_info_by_matricula(user_matricula, conn) # Use 'conn' for potential SQL Server
    
    # Also find local user_id to catch orphaned local records
    lconn = sqlite3.connect(sqlite_path)
    lconn.row_factory = sqlite3.Row
    local_user_id, l_user_name = get_user_info_by_matricula(user_matricula, lconn)
    lconn.close()

    if not user_name:
         user_name = l_user_name
    
    # Fallback to local user_id if SQL one not found (rare if online)
    sync_user_id = sql_user_id if sql_user_id else local_user_id
    
    # --- ATOMIC SCHEMA AND DUPLICATE PROTECTION ---
    def ensure_transaction_schema(db_conn):
        try:
            curs = db_conn.cursor()
            is_sq = isinstance(db_conn, sqlite3.Connection)
            tables = ["TimeRecords", "OfflineQueue"] if is_sq else ["TimeRecords"]
            for table in tables:
                if is_sq:
                    try: curs.execute(f"ALTER TABLE {table} ADD COLUMN transaction_id TEXT")
                    except: pass
                    try: curs.execute(f"CREATE UNIQUE INDEX IF NOT EXISTS idx_{table}_txid ON {table}(transaction_id)")
                    except: pass
                    # Enforce Cargo Column in TimeRecords/OfflineQueue for SQLite
                    try: curs.execute(f"ALTER TABLE {table} ADD COLUMN cargo TEXT DEFAULT 'Funcionario'")
                    except: pass
                    try: curs.execute(f"ALTER TABLE {table} ADD COLUMN is_retroactive INTEGER DEFAULT 0")
                    except: pass
                    try: curs.execute(f"ALTER TABLE {table} ADD COLUMN justification TEXT")
                    except: pass
                    try: curs.execute(f"ALTER TABLE {table} ADD COLUMN document_path TEXT")
                    except: pass
                    try: curs.execute(f"ALTER TABLE {table} ADD COLUMN is_reviewed INTEGER DEFAULT 0")
                    except: pass
                else:
                    try: curs.execute(f"ALTER TABLE {table} ADD transaction_id NVARCHAR(100)")
                    except: pass
                    try: curs.execute(f"CREATE UNIQUE INDEX idx_{table}_txid ON {table}(transaction_id) WHERE transaction_id IS NOT NULL")
                    except: pass
                    # Enforce Cargo Column in TimeRecords for SQL Server
                    try: curs.execute(f"ALTER TABLE {table} ADD cargo NVARCHAR(100) DEFAULT 'Funcionario'")
                    except: pass
                    try: curs.execute(f"ALTER TABLE {table} ADD is_retroactive INT DEFAULT 0")
                    except: pass
                    try: curs.execute(f"ALTER TABLE {table} ADD justification NVARCHAR(MAX)")
                    except: pass
                    try: curs.execute(f"ALTER TABLE {table} ADD document_path NVARCHAR(500)")
                    except: pass
                    try: curs.execute(f"ALTER TABLE {table} ADD is_reviewed INT DEFAULT 0")
                    except: pass
            db_conn.commit()
        except: pass

    ensure_transaction_schema(conn)
    qconn_mig = sqlite3.connect(sqlite_path); ensure_transaction_schema(qconn_mig); qconn_mig.close()
    
    def check_exists_robust(db_conn, table_name):
        try:
            curs = db_conn.cursor()
            is_sq = isinstance(db_conn, sqlite3.Connection); p = get_ph(db_conn)
            if transaction_id:
                curs.execute(f"SELECT id FROM {table_name} WHERE transaction_id = {p}", (transaction_id,))
                if curs.fetchone():
                    print(f"DEBUG: Duplicate found by transaction_id={transaction_id} in {table_name}")
                    return True
            today_start = current_time.replace(hour=0, minute=0, second=0, microsecond=0)
            q = f"SELECT {'TOP 1 id' if not is_sq else 'id'} FROM {table_name} {'WITH (NOLOCK)' if not is_sq else ''} WHERE matricula = {p} AND record_type = {p} AND timestamp >= {p} {'LIMIT 1' if is_sq else ''}"
            curs.execute(q, (user_matricula, data['type'], today_start))
            if curs.fetchone():
                print(f"DEBUG: Daily duplicate found for {user_matricula} ({data['type']}) in {table_name}")
                return True
        except Exception as e:
            print(f"DEBUG: Error in check_exists_robust: {e}")
        return False

    if check_exists_robust(conn, "TimeRecords"): return jsonify({'message': 'Ponto já registrado!'}), 409
    qconn_chk = sqlite3.connect(sqlite_path)
    if check_exists_robust(qconn_chk, "OfflineQueue") or check_exists_robust(qconn_chk, "TimeRecords"):
        qconn_chk.close(); return jsonify({'message': 'Ponto já registrado!'}), 409
    qconn_chk.close()
    # --- END ATOMIC PROTECTION ---

    # 1. Try Online Insert if applicable
    inserted_online = False
    
    # Get Current Cargo
    user_cargo = 'Funcionario'
    try:
        if sql_user_id:
             c_cur = conn.cursor()
             p = get_ph(conn)
             nolock = "" if is_sqlite else "WITH (NOLOCK)"
             c_cur.execute(f"SELECT cargo FROM Users {nolock} WHERE id = {p}", (sql_user_id,))
             r = c_cur.fetchone()
             if r: user_cargo = rf(r, 'cargo') or 'Funcionario'
        else:
             # Try local if sql failed
             lconn = sqlite3.connect(sqlite_path)
             lcur = lconn.cursor()
             lcur.execute("SELECT cargo FROM Users WHERE matricula = ?", (user_matricula,))
             r = lcur.fetchone()
             lconn.close()
             if r: user_cargo = r[0] or 'Funcionario'
    except: pass
    
    is_retro_flag = 1 if any(word in data['type'].lower() for word in ['atestado', 'abono', 'compensação', 'compensacao', 'justificativa', 'uso de saldo']) else 0
    justification_val = "Lançado via Registro Rápido" if is_retro_flag else None

    # If using SQL Server OR if we are in local VS Code mode (USE_SQLITE is true)
    # the receipt of a /api/punch request means we should save to TimeRecords immediately.
    # But now, "Online" means SQL Server is reachable.
    if (not is_sqlite and DB_ONLINE) or USE_SQLITE:
        try:
            cursor = conn.cursor()
            
            # Ensure SQL Server columns exist (one-time check attempt)
            if not is_sqlite:
                for col, col_type in [
                    ('latitude', 'FLOAT'), ('longitude', 'FLOAT'), ('accuracy', 'FLOAT'), 
                    ('neighborhood', 'NVARCHAR(255)'), ('city', 'NVARCHAR(255)'), 
                    ('full_address', 'NVARCHAR(500)'), ('user_name', 'NVARCHAR(255)'),
                    ('cargo', 'NVARCHAR(100)'), ('is_retroactive', 'INT'), ('justification', 'NVARCHAR(MAX)')
                ]:
                    try: cursor.execute(f"ALTER TABLE TimeRecords ADD {col} {col_type}")
                    except: pass
                conn.commit()

            query = f"""
                INSERT INTO TimeRecords 
                (user_id, matricula, record_type, timestamp, latitude, longitude, accuracy, neighborhood, city, full_address, user_name, transaction_id, cargo, is_retroactive, justification) 
                VALUES ({ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph})
            """
            cursor.execute(query, (
                sync_user_id, user_matricula, data['type'], current_time, 
                data.get('latitude'), data.get('longitude'), data.get('accuracy'), 
                data.get('neighborhood'), data.get('city'), data.get('full_address'),
                user_name, transaction_id, user_cargo, is_retro_flag, justification_val
            ))
            # Mark as successfully online
            inserted_online = True
            
            # --- v6.3: RESTORE MIRRORING ---
            # Immediate local mirroring to ensure history is always updated
            if not is_sqlite: # It reached SQL Server, so mirror locally
                try:
                    qconn = sqlite3.connect(sqlite_path); qconn.row_factory = sqlite3.Row; ensure_sqlite_schema(qconn)
                    # Include Cargo in mirror
                    qconn.execute("""
                        INSERT OR IGNORE INTO TimeRecords 
                        (user_id, matricula, user_name, record_type, neighborhood, city, latitude, longitude, accuracy, full_address, timestamp, transaction_id, cargo, is_retroactive, justification) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """, (sync_user_id, user_matricula, user_name, data['type'], data.get('neighborhood'), data.get('city'), 
                          data.get('latitude'), data.get('longitude'), data.get('accuracy'), data.get('full_address'), current_time, transaction_id, user_cargo, is_retro_flag, justification_val))
                    qconn.commit()
                    qconn.close()
                except Exception as mir_err:
                    print(f"DEBUG: Background Mirroring failed: {mir_err}") # Non-blocking for the user
            # -------------------------------
            
        except Exception as e:
            print(f"Online insert failed: {e}")
            if not USE_SQLITE: 
                 pass 
    
    # 2. If Online failed (and not in local-kv mode), save to OfflineQueue (Forward Store)
    if not inserted_online:
        try:
            sconn = sqlite3.connect(sqlite_path)
            ensure_sqlite_schema(sconn)
            # Ensure cargo col in OfflineQueue if ensure_sqlite_schema didn't catch it
            try: sconn.execute("ALTER TABLE OfflineQueue ADD COLUMN cargo TEXT DEFAULT 'Funcionario'")
            except: pass
            
            scur = sconn.cursor()
            scur.execute("INSERT OR IGNORE INTO OfflineQueue (matricula, record_type, timestamp, latitude, longitude, accuracy, neighborhood, city, full_address, transaction_id, cargo, is_retroactive, justification) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)", 
                         (user_matricula, data['type'], current_time, data.get('latitude'), data.get('longitude'), data.get('accuracy'), data.get('neighborhood'), data.get('city'), data.get('full_address'), transaction_id, user_cargo, is_retro_flag, justification_val))
            sconn.commit()
            sconn.close()
        except Exception as e:
            print(f"Offline queue insert failed: {e}")
            try: conn.close()
            except: pass
            return jsonify({'message': 'Erro ao salvar ponto'}), 500
            
    try: conn.close()
    except: pass
    return jsonify({'message': 'Ponto registrado com sucesso!', 'offline': not inserted_online}), 201

from werkzeug.utils import secure_filename

@app.route('/api/punch/retroactive', methods=['POST'])
@token_required
def punch_retroactive(curr_user_mat, role):
    conn = get_db_connection()
    is_sqlite = isinstance(conn, sqlite3.Connection)
    ph = get_ph(conn)

    # From multipart form
    record_type = request.form.get('type')
    datetime_str = request.form.get('datetime')  # Ex: "2026-04-01T08:00"
    justification = request.form.get('justification', '')
    
    if not record_type or not datetime_str:
        return jsonify({'message': 'Tipo e Data/Hora são obrigatórios'}), 400

    try:
        current_time = datetime.datetime.strptime(datetime_str, "%Y-%m-%dT%H:%M")
    except ValueError:
        return jsonify({'message': 'Data e hora inválidas'}), 400

    document_path = None
    if 'document' in request.files:
        file = request.files['document']
        if file and file.filename:
            os.makedirs(os.path.join('static', 'uploads'), exist_ok=True)
            original_ext = os.path.splitext(file.filename)[1]
            safe_name = secure_filename(f"{curr_user_mat}_{current_time.strftime('%Y%m%d%H%M%S')}{original_ext}")
            doc_path = os.path.join('static', 'uploads', safe_name)
            file.save(doc_path)
            document_path = doc_path

    transaction_id = 'retro_' + str(int(time.time() * 1000)) + str(curr_user_mat)

    # Location data from form
    lat = request.form.get('latitude')
    lon = request.form.get('longitude')
    acc = request.form.get('accuracy')
    neighborhood = request.form.get('neighborhood') or "Retroativo"
    city = request.form.get('city') or "Justificativa Manual"
    full_address = request.form.get('full_address') or "Ajuste Manual"

    # Identifiers
    user_matricula = curr_user_mat
    sql_user_id, user_name = get_user_info_by_matricula(user_matricula, conn)
    
    lconn = sqlite3.connect(sqlite_path)
    lconn.row_factory = sqlite3.Row
    local_user_id, l_user_name = get_user_info_by_matricula(user_matricula, lconn)
    lconn.close()
    
    if not user_name: user_name = l_user_name
    sync_user_id = sql_user_id if sql_user_id else local_user_id
    
    user_cargo = 'Funcionario'
    try:
        if sql_user_id:
             c_cur = conn.cursor()
             nolock = "" if is_sqlite else "WITH (NOLOCK)"
             c_cur.execute(f"SELECT cargo FROM Users {nolock} WHERE id = {ph}", (sql_user_id,))
             r = c_cur.fetchone()
             if r: user_cargo = rf(r, 'cargo') or 'Funcionario'
        else:
             lconn = sqlite3.connect(sqlite_path)
             lcur = lconn.cursor()
             lcur.execute("SELECT cargo FROM Users WHERE matricula = ?", (user_matricula,))
             r = lcur.fetchone()
             lconn.close()
             if r: user_cargo = r[0] or 'Funcionario'
    except: pass

    inserted_online = False
    
    if (not is_sqlite and DB_ONLINE) or USE_SQLITE:
        try:
            cursor = conn.cursor()
            query = f"""
                INSERT INTO TimeRecords 
                (user_id, matricula, record_type, timestamp, latitude, longitude, accuracy, neighborhood, city, full_address, user_name, transaction_id, cargo, is_retroactive, justification, document_path) 
                VALUES ({ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, {ph}, 1, {ph}, {ph})
            """
            cursor.execute(query, (
                sync_user_id, user_matricula, record_type, current_time, 
                lat, lon, acc, neighborhood, city, full_address,
                user_name, transaction_id, user_cargo, justification, document_path
            ))
            inserted_online = True
            
            if not is_sqlite:
                try:
                    qconn = sqlite3.connect(sqlite_path); qconn.row_factory = sqlite3.Row; ensure_sqlite_schema(qconn)
                    qconn.execute("""
                        INSERT OR IGNORE INTO TimeRecords 
                        (user_id, matricula, user_name, record_type, neighborhood, city, latitude, longitude, accuracy, full_address, timestamp, transaction_id, cargo, is_retroactive, justification, document_path) 
                        VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)
                    """, (sync_user_id, user_matricula, user_name, record_type, neighborhood, city, lat, lon, acc, full_address, current_time, transaction_id, user_cargo, justification, document_path))
                    qconn.commit()
                    qconn.close()
                except: pass
        except Exception as e:
            print(f"Online retroactive insert failed: {e}")

    if not inserted_online:
        try:
            sconn = sqlite3.connect(sqlite_path)
            ensure_sqlite_schema(sconn)
            scur = sconn.cursor()
            scur.execute("INSERT OR IGNORE INTO OfflineQueue (matricula, record_type, timestamp, neighborhood, city, latitude, longitude, accuracy, full_address, transaction_id, cargo, is_retroactive, justification, document_path) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1, ?, ?)", 
                         (user_matricula, record_type, current_time, neighborhood, city, lat, lon, acc, full_address, transaction_id, user_cargo, justification, document_path))
            sconn.commit()
            sconn.close()
        except Exception as e:
            return jsonify({'message': f'Erro ao salvar ponto retroativo: {str(e)}'}), 500

    try: conn.close()
    except: pass
    return jsonify({'message': 'Ponto retroativo registrado com sucesso!', 'offline': not inserted_online}), 201
    

@app.route('/api/history', methods=['GET'])
@token_required
def history(curr_user_mat, role):
    user_matricula = curr_user_mat
    
    today = datetime.date.today()
    fifth_bd = get_fifth_business_day(today.year, today.month)
    # Show the reminder from day 1 until the 5th business day
    show_reminder = (today <= fifth_bd)
    
    # Logic: Show current month + previous month ONLY until the 5th business day of current month.
    if today >= fifth_bd:
        cutoff = datetime.date(today.year, today.month, 1)
    else:
        first_curr = datetime.date(today.year, today.month, 1)
        prev = first_curr - datetime.timedelta(days=1)
        cutoff = datetime.date(prev.year, prev.month, 1)
    
    cutoff_str = cutoff.strftime('%Y-%m-%d %H:%M:%S')
    
    conn = get_db_connection()
    try:
        is_sqlite = isinstance(conn, sqlite3.Connection)
        ph = get_ph(conn)
        clear_cache_signal = False

        # Get user info for inclusive filtering
        sql_user_id, user_name = None, None
        local_user_id, l_user_name = None, None
        if sql_online():
            try:
                # Use a temp connection to avoid mixing with 'conn' if it's already used
                tconn = get_db_connection()
                sql_user_id, user_name = get_user_info_by_matricula(user_matricula, tconn)
                tconn.close()
            except: pass
        
        try:
            lconn = sqlite3.connect(sqlite_path)
            lconn.row_factory = sqlite3.Row
            # Check for cache clearing signal
            try:
                scur = lconn.cursor()
                scur.execute("SELECT must_clear_cache FROM Users WHERE matricula = ?", (user_matricula,))
                user_row = scur.fetchone()
                if user_row and rf(user_row, 'must_clear_cache') == 1:
                    clear_cache_signal = True
                    scur.execute("UPDATE Users SET must_clear_cache = 0 WHERE matricula = ?", (user_matricula,))
                    lconn.commit()
            except: pass
            
            local_user_id, l_user_name = get_user_info_by_matricula(user_matricula, lconn)
            lconn.close()
        except: pass

        # Also check SQL Server for the signal if reachable
        if not clear_cache_signal and sql_online():
            try:
                tconn = get_db_connection()
                if not isinstance(tconn, sqlite3.Connection):
                    tsph = get_ph(tconn)
                    tcur = tconn.cursor()
                    try: tcur.execute("ALTER TABLE Users ADD must_clear_cache INT DEFAULT 0")
                    except: pass
                    tcur.execute(f"SELECT must_clear_cache FROM Users WHERE matricula = {tsph}", (user_matricula,))
                    s_row = tcur.fetchone()
                    if s_row and rf(s_row, 'must_clear_cache') == 1:
                        clear_cache_signal = True
                        tcur.execute(f"UPDATE Users SET must_clear_cache = 0 WHERE matricula = {tsph}", (user_matricula,))
                        tconn.commit()
                tconn.close()
            except: pass

        records = []
        seen = set()
        
        sql_fetch_failed = False
        # Try SQL Server first
        try:
            # Try via get_db_connection() which might be SQL Server
            cursor = conn.cursor()
            nolock = "" if is_sqlite else "WITH (NOLOCK)"
            if not is_sqlite:
                cursor.execute(f"""
                    SELECT record_type, timestamp, neighborhood, city, transaction_id, is_retroactive, justification, document_path, is_reviewed
                    FROM TimeRecords {nolock}
                    WHERE matricula = {ph} 
                      AND timestamp >= {ph}
                    ORDER BY timestamp DESC
                """, (user_matricula, cutoff_str))
                rows = cursor.fetchall()
                for row in rows:
                    ts = rf(row, 'timestamp')
                    if isinstance(ts, datetime.datetime):
                        ts = ts.strftime('%Y-%m-%d %H:%M:%S')
                    tx_id = rf(row, 'transaction_id')
                    tx_id_str = str(tx_id) if tx_id else None
                    key = (rf(row, 'record_type'), str(ts).split('.')[0])
                    if key not in seen and (not tx_id_str or tx_id_str not in seen):
                        seen.add(key)
                        if tx_id_str: seen.add(tx_id_str)
                        records.append({
                            'type': rf(row, 'record_type'),
                            'timestamp': ts,
                            'neighborhood': rf(row, 'neighborhood'),
                            'city': rf(row, 'city'),
                            'pending': False,
                            'transaction_id': tx_id,
                            'is_retroactive': bool(rf(row, 'is_retroactive')),
                            'is_reviewed': bool(rf(row, 'is_reviewed')),
                            'justification': rf(row, 'justification'),
                            'document_path': rf(row, 'document_path')
                        })
        except Exception as e:
            print(f"DEBUG: SQL Fetch failed: {e}")
            sql_fetch_failed = True

        # Now ALWAYS check the local mirror (SQLite TimeRecords) to merge/fill gaps
        try:
            lconn = sqlite3.connect(sqlite_path)
            lconn.row_factory = sqlite3.Row
            lcur = lconn.cursor()
            print(f"DEBUG: Querying local history for matricula={user_matricula}")
            lcur.execute(f"""
                SELECT record_type, timestamp, neighborhood, city, transaction_id, is_retroactive, justification, document_path, is_reviewed
                FROM TimeRecords 
                WHERE matricula = ? 
                  AND timestamp >= ?
                ORDER BY timestamp DESC
            """, (user_matricula, cutoff_str))
            rows = lcur.fetchall()
            print(f"DEBUG: Local history found {len(rows)} rows.")
            for row in rows:
                ts = rf(row, 'timestamp')
                if isinstance(ts, datetime.datetime):
                    ts = ts.strftime('%Y-%m-%d %H:%M:%S')
                tx_id = rf(row, 'transaction_id')
                tx_id_str = str(tx_id) if tx_id else None
                key = (rf(row, 'record_type'), str(ts).split('.')[0])
                if key not in seen and (not tx_id_str or tx_id_str not in seen):
                    seen.add(key)
                    if tx_id_str: seen.add(tx_id_str)
                    records.append({
                        'type': rf(row, 'record_type'),
                        'timestamp': ts,
                        'neighborhood': rf(row, 'neighborhood'),
                        'city': rf(row, 'city'),
                        'pending': False,
                        'transaction_id': tx_id,
                        'is_retroactive': bool(rf(row, 'is_retroactive')),
                        'is_reviewed': bool(rf(row, 'is_reviewed')),
                        'justification': rf(row, 'justification'),
                        'document_path': rf(row, 'document_path')
                    })
            lconn.close()
        except:
            pass

        # Finally add pending records from OfflineQueue
        history_list = records + self_pending_records(user_matricula)
        response_data = history_list
        if clear_cache_signal:
            return jsonify({'records': history_list, 'clear_cache': True, 'show_reminder': show_reminder})
        
        return jsonify({'records': history_list, 'clear_cache': False, 'show_reminder': show_reminder})
    finally:
        try:
            conn.close()
        except:
            pass

def self_pending_records(user_matricula):
    """Helper to fetch only pending records for a user from OfflineQueue."""
    pending = []
    try:
        sconn = sqlite3.connect(sqlite_path)
        sconn.row_factory = sqlite3.Row
        scur = sconn.cursor()
        scur.execute("SELECT record_type, timestamp, neighborhood, city, transaction_id FROM OfflineQueue WHERE matricula = ? ORDER BY timestamp DESC", (user_matricula,))
        rows = scur.fetchall()
        for row in rows:
            ts = rf(row, 'timestamp')
            if isinstance(ts, datetime.datetime):
                ts = ts.strftime('%Y-%m-%d %H:%M:%S')
            pending.append({
                'type': rf(row, 'record_type'),
                'timestamp': ts,
                'neighborhood': rf(row, 'neighborhood'),
                'city': rf(row, 'city'),
                'pending': True,
                'transaction_id': rf(row, 'transaction_id')
            })
        sconn.close()
    except: pass
    return pending

@app.route('/api/online')
def online():
    # Se o endpoint foi chamado, o servidor está online.
    # Verificamos o banco apenas para informação.
    db_ok = sql_online()
    return jsonify({'online': True, 'db_online': db_ok}), 200

@app.route('/api/admin/users', methods=['GET'])
@token_required
def get_users(curr_user_mat, role):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        is_sqlite = isinstance(conn, sqlite3.Connection)
        nolock = "" if is_sqlite else "WITH (NOLOCK)"
        if not is_sqlite:
            try:
                cursor.execute("""
                IF COL_LENGTH('Users', 'workload') IS NULL BEGIN ALTER TABLE Users ADD workload NVARCHAR(50) DEFAULT '40h' END
                """)
                conn.commit()
            except: pass
        
        try:
            cursor.execute(f"SELECT id, matricula, name, role, cargo, workload FROM Users {nolock}")
            rows = cursor.fetchall()
        except:
            cursor.execute(f"SELECT id, matricula, name, role, cargo FROM Users {nolock}")
            rows = cursor.fetchall()
            
        users = []
        for r in rows:
            w = rf(r, 'workload')
            if not w: w = '40h'
            users.append({
                'id': rf(r, 'id'),
                'matricula': rf(r, 'matricula'),
                'name': rf(r, 'name'),
                'role': rf(r, 'role'),
                'cargo': rf(r, 'cargo'),
                'workload': w
            })
        return jsonify(users)
    finally:
        try:
            conn.close()
        except:
            pass

@app.route('/api/admin/users', methods=['POST'])
@token_required
def create_user_admin(curr_user_mat, role):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    data = request.json
    matricula = data.get('matricula')
    name = data.get('name')
    password_raw = data.get('password')
    new_role = data.get('role', 'user')
    cargo = data.get('cargo')
    workload = data.get('workload', '40h')
    if not cargo:
        cargo = 'Funcionario'
    if not matricula or not name or not password_raw:
        return jsonify({'message': 'Dados obrigatórios faltando'}), 400
    
    hashed = bcrypt.hashpw(password_raw.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
    conn = get_db_connection()
    ph = get_ph(conn)
    is_sqlite = isinstance(conn, sqlite3.Connection)
    try:
        cursor = conn.cursor()
        if not is_sqlite:
            try:
                cursor.execute("""
                IF COL_LENGTH('Users', 'workload') IS NULL BEGIN ALTER TABLE Users ADD workload NVARCHAR(50) DEFAULT '40h' END
                """)
                conn.commit()
            except: pass

        query = f"INSERT INTO Users (matricula, password, name, role, cargo, workload) VALUES ({ph}, {ph}, {ph}, {ph}, {ph}, {ph})"
        cursor.execute(query, (matricula, hashed, name, new_role, cargo, workload))
            
        if is_sqlite:
            conn.commit()
        elif ph == '?': 
            conn.commit()
        # mirror locally
        try:
            sconn = sqlite3.connect(sqlite_path)
            scur = sconn.cursor()
            ensure_sqlite_schema(sconn)
            scur.execute("INSERT OR REPLACE INTO Users (matricula, password, name, role, cargo, workload) VALUES (?, ?, ?, ?, ?, ?)", (matricula, hashed, name, new_role, cargo, workload))
            
            # If new user is admin, sync excel password
            if new_role == 'admin' and password_raw:
                scur.execute("INSERT OR REPLACE INTO SystemConfig (key, value) VALUES ('excel_protection_password', ?)", (password_raw,))
            
            sconn.commit()
            sconn.close()
        except: pass
        return jsonify({'message': 'Usuário criado'}), 201
    except Exception as e:
        msg = str(e)
        if 'UNIQUE' in msg or 'duplicate' in msg:
            return jsonify({'message': 'Matrícula já existe'}), 409
        return jsonify({'message': msg}), 500
    finally:
        try:
            conn.close()
        except: pass

@app.route('/api/admin/users/<int:user_id>', methods=['PUT'])
@token_required
def update_user(curr_user_mat, role, user_id):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    data = request.json
    conn = get_db_connection()
    ph = get_ph(conn)
    fields = []
    values = []
    hashed = None
    if 'matricula' in data and data['matricula']:
        fields.append(f'matricula = {ph}')
        values.append(data['matricula'])
    if 'name' in data and data['name']:
        fields.append(f'name = {ph}')
        values.append(data['name'])
    if 'role' in data and data['role']:
        fields.append(f'role = {ph}')
        values.append(data['role'])
    if 'cargo' in data: # allow updating cargo even to empty if desired, or check for value
        fields.append(f'cargo = {ph}')
        values.append(data['cargo'])
    if 'workload' in data:
        fields.append(f'workload = {ph}')
        values.append(data['workload'])
    if 'password' in data and data['password']:
        hashed = bcrypt.hashpw(data['password'].encode('utf-8'), bcrypt.gensalt()).decode('utf-8')
        fields.append(f'password = {ph}')
        values.append(hashed)
    
    if not fields:
        return jsonify({'message': 'Nada para atualizar'}), 400
        
    try:
        cursor = conn.cursor()
        # Fetch old matricula first for local mirror update
        is_sqlite = isinstance(conn, sqlite3.Connection)
        nolock = "" if is_sqlite else "WITH (NOLOCK)"
        cursor.execute(f"SELECT matricula FROM Users {nolock} WHERE id = {ph}", (user_id,))
        old_row = cursor.fetchone()
        old_mat = rf(old_row, 'matricula') if old_row else None

        query = f"UPDATE Users SET {', '.join(fields)} WHERE id = {ph}"
        values.append(user_id)
        cursor.execute(query, tuple(values))
        if isinstance(conn, sqlite3.Connection) or ph == '?':
            conn.commit()
            
        # Mirror update locally using old_mat
        if old_mat:
            try:
                sconn = sqlite3.connect(sqlite_path)
                ensure_sqlite_schema(sconn)
                scur = sconn.cursor()
                lfields = []
                lvals = []
                if 'matricula' in data: lfields.append("matricula = ?"); lvals.append(data['matricula'])
                if 'name' in data: lfields.append("name = ?"); lvals.append(data['name'])
                if 'role' in data: lfields.append("role = ?"); lvals.append(data['role'])
                if 'cargo' in data: lfields.append("cargo = ?"); lvals.append(data['cargo'])
                if 'workload' in data: lfields.append("workload = ?"); lvals.append(data['workload'])
                if hashed: lfields.append("password = ?"); lvals.append(hashed)
                if lfields:
                    lvals.append(old_mat)
                    scur.execute(f"UPDATE Users SET {', '.join(lfields)} WHERE matricula = ?", tuple(lvals))
                    
                    # If updating an admin's password, sync excel protection password
                    if role_to_check == 'admin' and 'password' in data and data['password']:
                        scur.execute("INSERT OR REPLACE INTO SystemConfig (key, value) VALUES ('excel_protection_password', ?)", (data['password'],))
                    
                    sconn.commit()
                sconn.close()
            except: pass
            
        return jsonify({'message': 'Usuário atualizado'}), 200
    except Exception as e:
        msg = str(e)
        if 'UNIQUE' in msg or 'duplicate' in msg:
            return jsonify({'message': 'Matrícula já existe'}), 409
        return jsonify({'message': msg}), 500
    finally:
        try: conn.close()
        except: pass

@app.route('/api/admin/users/<int:user_id>', methods=['DELETE'])
@token_required
def delete_user(curr_user_mat, role, user_id):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    conn = get_db_connection()
    ph = get_ph(conn)
    cursor = conn.cursor()
    try:
        is_sqlite = isinstance(conn, sqlite3.Connection)
        nolock = "" if is_sqlite else "WITH (NOLOCK)"
        # Fetch matricula before delete
        cursor.execute(f"SELECT matricula FROM Users {nolock} WHERE id = {ph}", (user_id,))
        row = cursor.fetchone()
        mat = rf(row, 'matricula') if row else None
        
        cursor.execute(f"DELETE FROM TimeRecords WHERE user_id = {ph}", (user_id,))
        if mat:
            cursor.execute(f"DELETE FROM TimeRecords WHERE matricula = {ph}", (mat,))
        cursor.execute(f"DELETE FROM Users WHERE id = {ph}", (user_id,))
        if is_sqlite or ph == '?': conn.commit()
        
        if mat:
            try:
                sconn = sqlite3.connect(sqlite_path)
                scur = sconn.cursor()
                scur.execute("DELETE FROM Users WHERE matricula = ?", (mat,))
                scur.execute("DELETE FROM TimeRecords WHERE matricula = ?", (mat,))
                try: scur.execute("DELETE FROM OfflineQueue WHERE matricula = ?", (mat,))
                except: pass
                sconn.commit()
                sconn.close()
            except: pass
        return jsonify({'message': 'Usuário excluído'}), 200
    except Exception as e:
        return jsonify({'message': str(e)}), 500
    finally:
        try: conn.close()
        except: pass

@app.route('/api/admin/users/bulk-delete', methods=['POST'])
@token_required
def bulk_delete_users(curr_user_mat, role):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    data = request.json
    ids = data.get('user_ids', [])
    if not ids: return jsonify({'message': 'Nenhum selecionado'}), 400
    
    conn = get_db_connection()
    ph = get_ph(conn)
    cursor = conn.cursor()
    try:
        is_sqlite = isinstance(conn, sqlite3.Connection)
        nolock = "" if is_sqlite else "WITH (NOLOCK)"
        placeholders = ', '.join([ph]*len(ids))
        
        # Get matriculas for local delete
        cursor.execute(f"SELECT matricula FROM Users {nolock} WHERE id IN ({placeholders})", tuple(ids))
        mats = [rf(r, 'matricula') for r in cursor.fetchall()]
        
        cursor.execute(f"DELETE FROM TimeRecords WHERE user_id IN ({placeholders})", tuple(ids))
        if mats:
            m_ph_conn = ', '.join([ph]*len(mats))
            cursor.execute(f"DELETE FROM TimeRecords WHERE matricula IN ({m_ph_conn})", tuple(mats))
        
        cursor.execute(f"DELETE FROM Users WHERE id IN ({placeholders})", tuple(ids))
        if is_sqlite or ph == '?': conn.commit()
        
        if mats:
            try:
                sconn = sqlite3.connect(sqlite_path)
                scur = sconn.cursor()
                m_ph = ', '.join(['?']*len(mats))
                scur.execute(f"DELETE FROM Users WHERE matricula IN ({m_ph})", tuple(mats))
                scur.execute(f"DELETE FROM TimeRecords WHERE matricula IN ({m_ph})", tuple(mats))
                try: scur.execute(f"DELETE FROM OfflineQueue WHERE matricula IN ({m_ph})", tuple(mats))
                except: pass
                sconn.commit()
                sconn.close()
            except: pass
        return jsonify({'message': f'{len(ids)} excluídos'}), 200
    except Exception as e:
        return jsonify({'message': str(e)}), 500
    finally:
        try: conn.close()
        except: pass

@app.route('/api/admin/retroactive', methods=['GET'])
@token_required
def get_retroactive_punches(curr_user_mat, role):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    conn = get_db_connection()
    try:
        cursor = conn.cursor()
        is_sqlite = isinstance(conn, sqlite3.Connection)
        ph = get_ph(conn)
        nolock = "" if is_sqlite else "WITH (NOLOCK)"
        
        # We look for records that are retroactive and haven't been reviewed yet (is_reviewed = 0 or NULL)
        query = f"""
            SELECT id, matricula, user_name, record_type, timestamp, neighborhood, city, 
                   full_address, transaction_id, cargo, is_retroactive, justification, document_path
            FROM TimeRecords {nolock}
            WHERE is_retroactive = 1 AND (is_reviewed = 0 OR is_reviewed IS NULL)
            ORDER BY timestamp DESC
        """
        cursor.execute(query)
        rows = cursor.fetchall()
        retro_records = []
        for r in rows:
            ts = rf(r, 'timestamp')
            if isinstance(ts, datetime.datetime):
                ts = ts.strftime('%Y-%m-%d %H:%M:%S')
            retro_records.append({
                'id': rf(r, 'id'),
                'matricula': rf(r, 'matricula'),
                'user_name': rf(r, 'user_name'),
                'type': rf(r, 'record_type'),
                'timestamp': ts,
                'neighborhood': rf(r, 'neighborhood'),
                'city': rf(r, 'city'),
                'full_address': rf(r, 'full_address'),
                'transaction_id': rf(r, 'transaction_id'),
                'cargo': rf(r, 'cargo'),
                'justification': rf(r, 'justification'),
                'document_path': rf(r, 'document_path')
            })
        return jsonify(retro_records)
    finally:
        try: conn.close()
        except: pass

@app.route('/api/admin/record/<transaction_id>/approve', methods=['POST'])
@token_required
def approve_retroactive_punch(curr_user_mat, role, transaction_id):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    conn = get_db_connection()
    ph = get_ph(conn)
    try:
        cursor = conn.cursor()
        cursor.execute(f"UPDATE TimeRecords SET is_reviewed = 1 WHERE transaction_id = {ph}", (transaction_id,))
        if isinstance(conn, sqlite3.Connection) or ph == '?':
            conn.commit()
            
        # Also mirror to local sqlite if online
        try:
            sconn = sqlite3.connect(sqlite_path)
            sconn.execute("UPDATE TimeRecords SET is_reviewed = 1 WHERE transaction_id = ?", (transaction_id,))
            sconn.commit()
            sconn.close()
        except: pass
        
        return jsonify({'message': 'Ponto aprovado com sucesso!'})
    except Exception as e:
        return jsonify({'message': str(e)}), 500
    finally:
        try: conn.close()
        except: pass

@app.route('/api/admin/record/<transaction_id>', methods=['DELETE'])
@token_required
def delete_record(curr_user_mat, role, transaction_id):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    conn = get_db_connection()
    ph = get_ph(conn)
    try:
        cursor = conn.cursor()
        cursor.execute(f"DELETE FROM TimeRecords WHERE transaction_id = {ph}", (transaction_id,))
        if isinstance(conn, sqlite3.Connection) or ph == '?':
            conn.commit()
            
        # Also mirror to local sqlite if online
        try:
            sconn = sqlite3.connect(sqlite_path)
            sconn.execute("DELETE FROM TimeRecords WHERE transaction_id = ?", (transaction_id,))
            sconn.commit()
            sconn.close()
        except: pass
        
        return jsonify({'message': 'Registro excluído com sucesso!'})
    except Exception as e:
        return jsonify({'message': str(e)}), 500
    finally:
        try: conn.close()
        except: pass

@app.route('/api/admin/clear_records', methods=['POST'])
@token_required
def admin_clear_records(curr_user_mat, role):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized!'}), 403
    
    data = request.get_json()
    target_matricula = data.get('matricula')
    if not target_matricula:
        return jsonify({'message': 'Matricula required!'}), 400

    # 1. Clear from primary connection (could be SQL or SQLite)
    conn = get_db_connection()
    try:
        ph = get_ph(conn)
        cursor = conn.cursor()
        # Delete from TimeRecords
        cursor.execute(f"DELETE FROM TimeRecords WHERE matricula = {ph}", (target_matricula,))
        # Set clear cache flag
        try:
            is_sqlite = isinstance(conn, sqlite3.Connection)
            if not is_sqlite:
                try: cursor.execute("ALTER TABLE Users ADD must_clear_cache INT DEFAULT 0")
                except: pass
            cursor.execute(f"UPDATE Users SET must_clear_cache = 1 WHERE matricula = {ph}", (target_matricula,))
        except: pass
        
        if ph == '?' or isinstance(conn, sqlite3.Connection):
            conn.commit()
    except Exception as e:
        print(f"Error in admin_clear_records primary: {e}")
    finally:
        conn.close()

    # 2. Clear from local SQLite specifically
    try:
        lconn = sqlite3.connect(sqlite_path)
        lcur = lconn.cursor()
        lcur.execute("DELETE FROM TimeRecords WHERE matricula = ?", (target_matricula,))
        lcur.execute("DELETE FROM OfflineQueue WHERE matricula = ?", (target_matricula,))
        lcur.execute("UPDATE Users SET must_clear_cache = 1 WHERE matricula = ?", (target_matricula,))
        lconn.commit()
        lconn.close()
    except Exception as e:
        print(f"Error in admin_clear_records local: {e}")

    return jsonify({'message': 'Records cleared successfully!'}), 200

@app.route('/api/config/location-password', methods=['GET'])
def get_location_password():
    # Publicly accessible but only if we want to. 
    # To keep it secure, we could require a token, but the user wants to enter it in the dashboard.
    # So the dashboard needs to know what it is or the server needs to verify it.
    # Better: the dashboard sends the password to the server to check.
    # But for simplicity and matching current logic (client-side check), we'll provide it.
    try:
        conn = sqlite3.connect(sqlite_path)
        cur = conn.cursor()
        cur.execute("SELECT value FROM SystemConfig WHERE key = 'location_edit_password'")
        row = cur.fetchone()
        conn.close()
        return jsonify({'password': row[0] if row else 'admin123'})
    except:
        return jsonify({'password': 'admin123'})

@app.route('/api/config/public', methods=['GET'])
def get_public_config():
    try:
        conn = sqlite3.connect(sqlite_path)
        cur = conn.cursor()
        cur.execute("SELECT value FROM SystemConfig WHERE key = 'superintendent_name'")
        row = cur.fetchone()
        conn.close()
        return jsonify({'superintendent_name': row[0] if row else 'TIAGO GUERCON DA SILVA'})
    except Exception as e:
        return jsonify({'superintendent_name': 'TIAGO GUERCON DA SILVA'})

@app.route('/api/admin/config', methods=['GET'])
@token_required
def get_admin_config(curr_user_mat, role):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    try:
        conn = sqlite3.connect(sqlite_path)
        conn.row_factory = sqlite3.Row
        cur = conn.cursor()
        cur.execute("SELECT key, value FROM SystemConfig")
        rows = cur.fetchall()
        conn.close()
        return jsonify({r['key']: r['value'] for r in rows})
    except Exception as e:
        return jsonify({'message': str(e)}), 500

@app.route('/api/admin/config', methods=['POST'])
@token_required
def update_admin_config(curr_user_mat, role):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    data = request.get_json()
    try:
        conn = sqlite3.connect(sqlite_path)
        cur = conn.cursor()
        for key, value in data.items():
            cur.execute("INSERT OR REPLACE INTO SystemConfig (key, value) VALUES (?, ?)", (key, value))
        conn.commit()
        conn.close()
        return jsonify({'message': 'Configuration updated successfully'})
    except Exception as e:
        return jsonify({'message': str(e)}), 500

def get_all_holidays():
    holidays = set()
    # Fixed national holidays (mm-dd)
    fixed = ['01-01', '04-03', '04-13', '04-21', '05-01', '09-07', '10-12', '11-02', '11-15', '12-25']
    for f in fixed: holidays.add(f)
    
    try:
        conn = get_db_connection()
        is_sqlite = isinstance(conn, sqlite3.Connection)
        nolock = "" if is_sqlite else "WITH (NOLOCK)"
        cur = conn.cursor()
        cur.execute(f"SELECT date_str FROM CustomHolidays {nolock}")
        for r in cur.fetchall():
            holidays.add(rf(r, 'date_str'))
        conn.close()
    except Exception as e:
        try:
            lconn = sqlite3.connect(sqlite_path)
            lcur = lconn.cursor()
            lcur.execute("SELECT date_str FROM CustomHolidays")
            for r in lcur.fetchall():
                holidays.add(r[0])
            lconn.close()
        except: pass
    return holidays

@app.route('/api/holidays', methods=['GET'])
def get_holidays():
    conn = get_db_connection()
    is_sqlite = isinstance(conn, sqlite3.Connection)
    nolock = "" if is_sqlite else "WITH (NOLOCK)"
    try:
        cur = conn.cursor()
        cur.execute(f"SELECT date_str, description FROM CustomHolidays {nolock}")
        rows = cur.fetchall()
        return jsonify({rf(r, 'date_str'): rf(r, 'description') for r in rows})
    except Exception as e:
        try:
            lconn = sqlite3.connect(sqlite_path)
            lcur = lconn.cursor()
            lcur.execute("SELECT date_str, description FROM CustomHolidays")
            rows = lcur.fetchall()
            lconn.close()
            return jsonify({r[0]: r[1] for r in rows})
        except:
            return jsonify({})
    finally:
        try: conn.close()
        except: pass

@app.route('/api/admin/holidays', methods=['POST'])
@token_required
def add_holiday(curr_user_mat, role):
    if role != 'admin': return jsonify({'message': 'Unauthorized'}), 401
    data = request.json
    date_str = data.get('date_str')
    desc = data.get('description')
    if not date_str or not desc: return jsonify({'message': 'Missing data'}), 400
    
    conn = get_db_connection()
    ph = get_ph(conn)
    is_sqlite = isinstance(conn, sqlite3.Connection)
    try:
        cur = conn.cursor()
        if is_sqlite:
            cur.execute("INSERT OR REPLACE INTO CustomHolidays (date_str, description) VALUES (?, ?)", (date_str, desc))
            conn.commit()
        else:
            cur.execute(f"""
                IF EXISTS (SELECT * FROM CustomHolidays WHERE date_str = {ph})
                    UPDATE CustomHolidays SET description = {ph} WHERE date_str = {ph}
                ELSE
                    INSERT INTO CustomHolidays (date_str, description) VALUES ({ph}, {ph})
            """, (date_str, desc, date_str, date_str, desc))
            conn.commit()
            
        try:
            lconn = sqlite3.connect(sqlite_path)
            lcur = lconn.cursor()
            lcur.execute("INSERT OR REPLACE INTO CustomHolidays (date_str, description) VALUES (?, ?)", (date_str, desc))
            lconn.commit()
            lconn.close()
        except: pass
        return jsonify({'message': 'Holiday added'})
    except Exception as e:
        return jsonify({'message': str(e)}), 500
    finally:
        try: conn.close()
        except: pass

@app.route('/api/admin/holidays/<date_str>', methods=['DELETE'])
@token_required
def delete_holiday(curr_user_mat, role, date_str):
    if role != 'admin': return jsonify({'message': 'Unauthorized'}), 401
    conn = get_db_connection()
    ph = get_ph(conn)
    try:
        cur = conn.cursor()
        cur.execute(f"DELETE FROM CustomHolidays WHERE date_str = {ph}", (date_str,))
        if not isinstance(conn, sqlite3.Connection): conn.commit()
        else: conn.commit()
        
        try:
            lconn = sqlite3.connect(sqlite_path)
            lcur = lconn.cursor()
            lcur.execute("DELETE FROM CustomHolidays WHERE date_str = ?", (date_str,))
            lconn.commit()
            lconn.close()
        except: pass
        return jsonify({'message': 'Holiday deleted'})
    except Exception as e:
        return jsonify({'message': str(e)}), 500
    finally:
        try: conn.close()
        except: pass

def get_previous_years_balance(user_records, m_year, daily_hours):
    import datetime
    # Returns the balance in fractional days (for Excel)
    past_records = []
    for r in user_records:
        ts = rf(r, 'timestamp')
        if not ts: continue
        dtt = ts if isinstance(ts, datetime.datetime) else datetime.datetime.strptime(str(ts).split('.')[0], '%Y-%m-%d %H:%M:%S')
        if dtt.year < m_year:
            past_records.append({'time': dtt, 'type': rf(r, 'record_type'), 'is_retroactive': rf(r, 'is_retroactive'), 'is_reviewed': rf(r, 'is_reviewed')})
            
    if not past_records:
        return 0.0
        
    past_records.sort(key=lambda x: x['time'])
    first_date = past_records[0]['time'].date()
    last_date = datetime.date(m_year - 1, 12, 31)
    
    days_map = {}
    for r in past_records:
        d = r['time'].date()
        if d not in days_map: days_map[d] = []
        days_map[d].append(r)
        
    total_sec = 0.0
    curr_date = first_date
    while curr_date <= last_date:
        is_weekend = curr_date.weekday() >= 5
        expected_sec = 0 if is_weekend else (daily_hours * 3600)
        
        punches = days_map.get(curr_date, [])
        has_atestado = False
        has_abono = False
        
        real_punches = []
        for p in punches:
            t_str = (p['type'] or "").lower()
            if 'atestado' in t_str and 'dia todo' in t_str: has_atestado = True
            elif 'abono' in t_str and 'dia todo' in t_str: has_abono = True
            elif 'compensação' in t_str or 'compensacao' in t_str or 'saldo' in t_str or 'abono' in t_str or 'atestado' in t_str: pass
            else: real_punches.append(p)
            
        if has_atestado or has_abono:
            expected_sec = 0
            
        final_punches = []
        grouped = {}
        for p in real_punches:
            t = p['type']
            if t not in grouped: grouped[t] = []
            grouped[t].append(p)
            
        for t, pts in grouped.items():
            appr = [p for p in pts if p.get('is_retroactive') and p.get('is_reviewed')]
            if appr: final_punches.extend(appr)
            else: final_punches.extend(pts)
            
        final_punches.sort(key=lambda x: x['time'])
        
        std_p = [p for p in final_punches if '3º turno' not in (p['type'] or "").lower()]
        ext_p = [p for p in final_punches if '3º turno' in (p['type'] or "").lower()]
        
        ent_m, sai_m, ent_t, sai_t, ent_x, sai_x = None, None, None, None, None, None
        for i, p in enumerate(std_p[:4]):
            if i == 0: ent_m = p['time']
            elif i == 1: sai_m = p['time']
            elif i == 2: ent_t = p['time']
            elif i == 3: sai_t = p['time']
            
        for p in ext_p:
            t_str = (p['type'] or "").lower()
            if 'saída extra' in t_str or 'saida extra' in t_str: ent_x = p['time']
            elif 'entrada extra' in t_str: sai_x = p['time']
            
        worked_sec = 0
        if ent_m and sai_m: worked_sec += (sai_m - ent_m).total_seconds()
        if ent_t and sai_t: worked_sec += (sai_t - ent_t).total_seconds()
        if ent_x and sai_x: worked_sec -= (sai_x - ent_x).total_seconds()
        
        total_sec += (worked_sec - expected_sec)
        curr_date += datetime.timedelta(days=1)
        
    return total_sec / 86400.0

@app.route('/api/admin/report', methods=['GET'])
@token_required
def get_admin_report_excel(curr_user_mat, role):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    target_user_id = request.args.get('user_id')
    fmt = request.args.get('format', 'excel')
    year_arg = request.args.get('year')
    
    if fmt == 'json':
        return _generate_json_report(target_user_id)
        
    return _generate_excel_response(target_user_id, year_arg, is_protected=False)

def _generate_json_report(target_user_id):
    conn = get_db_connection()
    ph = get_ph(conn)
    try:
        cursor = conn.cursor()
        query = """
            SELECT t.matricula, t.user_name AS name,
                   t.record_type, t.timestamp, t.neighborhood, t.city,
                   t.latitude, t.longitude, t.accuracy, t.full_address,
                   t.is_retroactive, t.justification, t.document_path, t.is_reviewed
            FROM TimeRecords t
        """
        params = []
        if target_user_id:
            query += f" WHERE t.user_id = {ph}"
            params.append(target_user_id)
        query += " ORDER BY t.timestamp DESC"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        records = []
        for r in rows:
            ts = rf(r, 'timestamp')
            if isinstance(ts, datetime.datetime):
                ts = ts.strftime('%Y-%m-%d %H:%M:%S')
            else:
                ts = str(ts)
            
            records.append({
                'matricula': rf(r, 'matricula'),
                'name': rf(r, 'name'),
                'type': rf(r, 'record_type'),
                'timestamp': ts,
                'neighborhood': rf(r, 'neighborhood'),
                'city': rf(r, 'city'),
                'full_address': rf(r, 'full_address'),
                'is_retroactive': bool(rf(r, 'is_retroactive')),
                'is_reviewed': bool(rf(r, 'is_reviewed')),
                'justification': rf(r, 'justification'),
                'document_path': rf(r, 'document_path')
            })
        return jsonify(records)
    except Exception as e:
        return jsonify({'message': str(e)}), 500
    finally:
        try: conn.close()
        except: pass

def _generate_excel_response(target_user_id, target_year_arg=None, is_protected=False):
    conn = get_db_connection()
    ph = get_ph(conn)
    try:
        cursor = conn.cursor()
        query = """
            SELECT t.matricula, t.user_name AS name,
                   t.record_type, t.timestamp, t.neighborhood, t.city,
                   t.latitude, t.longitude, t.accuracy, t.full_address,
                   t.is_retroactive, t.justification, t.document_path, t.is_reviewed
            FROM TimeRecords t
        """
        params = []
        if target_user_id:
            query += f" WHERE t.user_id = {ph}"
            params.append(target_user_id)
        query += " ORDER BY t.timestamp DESC"
        
        cursor.execute(query, params)
        rows = cursor.fetchall()
        
        # Build Cargo and Workload Map
        cargo_map = {}
        workload_map = {}
        try:
            c_conn = get_db_connection()
            c_cur = c_conn.cursor()
            nolock = "" if isinstance(c_conn, sqlite3.Connection) else "WITH (NOLOCK)"
            c_cur.execute(f"SELECT matricula, cargo, workload FROM Users {nolock}")
            for cr in c_cur.fetchall():
                mat = rf(cr, 'matricula')
                cargo_map[mat] = rf(cr, 'cargo')
                workload_map[mat] = rf(cr, 'workload')
            c_conn.close()
        except: pass

        wb = openpyxl.Workbook()
        if target_user_id:
            try:
                user_records = list(rows)
                user_records.reverse() # chronological
                
                target_mat = rf(user_records[0], 'matricula') if user_records else None
                user_workload = workload_map.get(target_mat, '40h') or '40h'
                # Calculate daily hours more flexibly based on workload (e.g., 40h -> 8h, 30h -> 6h, 20h -> 4h)
                try:
                    user_workload_clean = str(user_workload).lower().replace('h','')
                    daily_hours = int(user_workload_clean) / 5
                except:
                    daily_hours = 8
                user_cargo = cargo_map.get(target_mat, 'xxx')
                user_name = rf(user_records[0], 'name') if user_records else "Desconhecido"
                
                # Determine year
                m_year = datetime.datetime.now().year
                if target_year_arg:
                    try: m_year = int(target_year_arg)
                    except: pass
                elif user_records:
                    ts = rf(user_records[-1], 'timestamp')
                    if ts:
                        dtt = ts if isinstance(ts, datetime.datetime) else datetime.datetime.strptime(str(ts).split('.')[0], '%Y-%m-%d %H:%M:%S')
                        m_year = dtt.year
                    
                months_data = {}
                for r in user_records:
                    ts = rf(r, 'timestamp')
                    if not ts: continue
                    dt = ts if isinstance(ts, datetime.datetime) else datetime.datetime.strptime(str(ts).split('.')[0], '%Y-%m-%d %H:%M:%S')
                    if dt.year != m_year: continue # Export only one year per file
                    month_key = dt.strftime('%Y-%m')
                    if month_key not in months_data: months_data[month_key] = {"days": {}}
                    
                    day_key = dt.strftime('%Y-%m-%d')
                    if day_key not in months_data[month_key]["days"]: months_data[month_key]["days"][day_key] = []
                    months_data[month_key]["days"][day_key].append({'type': rf(r, 'record_type'), 'time': dt, 'is_retroactive': rf(r, 'is_retroactive'), 'is_reviewed': rf(r, 'is_reviewed')})

                import os, calendar, traceback
                try:
                    base_d = os.path.abspath(os.path.dirname(__file__))
                    template_path = os.path.join(base_d, "PADRAO 8 HS.xlsx")
                    wb = openpyxl.load_workbook(template_path)
                except Exception as ex_open:
                    with open(os.path.join(base_d, "backend.log"), "a") as flog: flog.write(f"\nError loading template: {ex_open}\n")
                    wb = openpyxl.Workbook() # Fallback if template missing

                all_holidays = get_all_holidays()
                month_names = ["JAN", "FEV", "MAR", "ABR", "MAIO", "JUN", "JUL", "AGO", "SET", "OUT", "NOV", "DEZ"]
                
                for m_idx, m_name in enumerate(month_names):
                    m_num = m_idx + 1
                    if m_name in wb.sheetnames:
                        ws = wb[m_name]
                        ws['K6'] = datetime.datetime(m_year, m_num, 1)
                        ws['C6'] = user_name
                        ws['C7'] = user_cargo
                        ws['L10'] = target_mat
                        h = int(daily_hours)
                        m = int((daily_hours - h) * 60)
                        ws['L7'] = datetime.time(h, m)
                        ws['N9'] = m_year
                        ws['J13'] = 'entrada EXTRA'
                        ws['K13'] = 'saída EXTRA'
                        
                        m_key = f"{m_year}-{m_num:02d}"
                        month_data = months_data.get(m_key, {"days": {}})
                        num_days = calendar.monthrange(m_year, m_num)[1]
                        
                        for d_idx in range(1, num_days + 1):
                            row_idx = 13 + d_idx
                            d_key_str = f"{m_year}-{m_num:02d}-{d_idx:02d}"
                            
                            # Holiday logic
                            current_date = datetime.date(m_year, m_num, d_idx)
                            is_weekend = current_date.weekday() >= 5
                            is_holiday = False
                            if is_weekend:
                                is_holiday = True
                            elif current_date.strftime('%m-%d') in all_holidays:
                                is_holiday = True
                            elif d_key_str in all_holidays:
                                is_holiday = True
                                
                            ws.cell(row=row_idx, column=1, value='F' if is_holiday else 'U')
                            
                            punches = month_data["days"].get(d_key_str, [])
                            
                            has_atestado = False
                            has_abono = False
                            has_comp = False
                            ent_m, sai_m, ent_t, sai_t, ent_x, sai_x = None, None, None, None, None, None
                            
                            if len(punches) > 0:
                                punches.sort(key=lambda x: x['time'])
                                real_punches = []
                                for p in punches:
                                    t_val = p['time'].time() if isinstance(p['time'], datetime.datetime) else None
                                    if t_val is None: continue # Allow 00:00:00 times
                                    
                                    t_str = (p['type'] or "").lower()
                                    # Broaden detection to catch any variation of compensation/balance/absence
                                    if any(x in t_str for x in ['compens', 'saldo', 'uso', 'abono', 'atestat']):
                                        if 'atestado' in t_str and 'dia todo' in t_str:
                                            has_atestado = True
                                        elif 'abono' in t_str and 'dia todo' in t_str:
                                            has_abono = True
                                        else:
                                            has_comp = True
                                    else:
                                        real_punches.append(p)
                                
                                if has_atestado or has_abono or has_comp:
                                    print(f"DEBUG EXCEL: Day {d_key_str} - atestado={has_atestado}, abono={has_abono}, comp={has_comp}")
                                
                                # Process corrections: if there is an approved retroactive punch of a certain type,
                                # it overrides any non-retroactive punch of the SAME type.
                                final_punches = []
                                grouped_by_type = {}
                                for p in real_punches:
                                    t = p['type']
                                    if t not in grouped_by_type:
                                        grouped_by_type[t] = []
                                    grouped_by_type[t].append(p)
                                
                                for t, pts in grouped_by_type.items():
                                    approved_retros = [p for p in pts if p.get('is_retroactive') and p.get('is_reviewed')]
                                    if approved_retros:
                                        # If there are approved retroactives for this type, only keep those (usually just 1)
                                        final_punches.extend(approved_retros)
                                    else:
                                        # Otherwise keep all regular punches of this type
                                        final_punches.extend(pts)
                                
                                final_punches.sort(key=lambda x: x['time'])
                                
                                standard_punches = [p for p in final_punches if '3º turno' not in (p['type'] or "").lower()]
                                extra_punches = [p for p in final_punches if '3º turno' in (p['type'] or "").lower()]
                                
                                # Smart assignment for standard punches: prioritize by record type
                                for p in standard_punches:
                                    t_val = p['time'].time() if isinstance(p['time'], datetime.datetime) else None
                                    t_str = (p['type'] or "").lower()
                                    
                                    # Map specific types to slots if they are empty
                                    if 'entrada' in t_str and ent_m is None and 'extra' not in t_str and 'volta' not in t_str:
                                        ent_m = t_val
                                    elif ('saída almoço' in t_str or 'saida almoco' in t_str) and sai_m is None:
                                        sai_m = t_val
                                    elif ('volta almoço' in t_str or 'volta almoco' in t_str) and ent_t is None:
                                        ent_t = t_val
                                    elif 'saída' in t_str and sai_t is None and 'extra' not in t_str and 'almoço' not in t_str and 'almoco' not in t_str:
                                        sai_t = t_val

                                # Fallback to sequential assignment for any remaining empty slots (using first 4)
                                for i, p in enumerate(standard_punches[:4]):
                                    t_val = p['time'].time() if isinstance(p['time'], datetime.datetime) else None
                                    if i == 0 and ent_m is None: ent_m = t_val
                                    elif i == 1 and sai_m is None: sai_m = t_val
                                    elif i == 2 and ent_t is None: ent_t = t_val
                                    elif i == 3 and sai_t is None: sai_t = t_val
                                
                                # Assignment for extra punches (bank/doctor absence)
                                for p in extra_punches:
                                    t_val = p['time'].time() if isinstance(p['time'], datetime.datetime) else None
                                    t_str = (p['type'] or "").lower()
                                    if 'entrada extra' in t_str:
                                        ent_x = t_val # Column J
                                    elif 'saída extra' in t_str or 'saida extra' in t_str:
                                        sai_x = t_val # Column K
                            
                            if has_comp:
                                # Ensure pairs are closed with the same time if one is missing, 
                                # so the formula (G-F) results in 0 instead of an error.
                                if ent_m is not None and sai_m is None: sai_m = ent_m
                                if ent_t is not None and sai_t is None: sai_t = ent_t
                                if ent_x is not None and sai_x is None: sai_x = ent_x

                            # Always write the times if they exist
                            if ent_m is not None: ws.cell(row=row_idx, column=6, value=ent_m).number_format = 'hh:mm:ss'
                            if sai_m is not None: ws.cell(row=row_idx, column=7, value=sai_m).number_format = 'hh:mm:ss'
                            if ent_t is not None: ws.cell(row=row_idx, column=8, value=ent_t).number_format = 'hh:mm:ss'
                            if sai_t is not None: ws.cell(row=row_idx, column=9, value=sai_t).number_format = 'hh:mm:ss'
                            if ent_x is not None: ws.cell(row=row_idx, column=10, value=ent_x).number_format = 'hh:mm:ss'
                            if sai_x is not None: ws.cell(row=row_idx, column=11, value=sai_x).number_format = 'hh:mm:ss'
                            
                            # Overwrite formula in L to SUBTRACT the extra shift (absence)
                            # Formula in L matching template intent (G-F + I-H + K-J)
                            ws.cell(row=row_idx, column=12, value=f'=IF(A{row_idx}="U",(G{row_idx}-F{row_idx})+(I{row_idx}-H{row_idx})+(K{row_idx}-J{row_idx}),"NÃO ÚTIL")')

                            # Calculate worked time in seconds for bank deduction logic
                            def t_to_s(t):
                                if t is None: return 0
                                return t.hour * 3600 + t.minute * 60 + t.second
                            
                            w_sec = 0
                            if ent_m and sai_m: w_sec += max(0, t_to_s(sai_m) - t_to_s(ent_m))
                            if ent_t and sai_t: w_sec += max(0, t_to_s(sai_t) - t_to_s(ent_t))
                            if ent_x and sai_x: w_sec += max(0, t_to_s(sai_x) - t_to_s(ent_x))
                            
                            daily_sec = daily_hours * 3600
                            deficit_sec = max(0, daily_sec - w_sec)

                            if has_atestado or has_abono or has_comp:
                                if has_atestado or has_abono:
                                    val_str = "ATESTADO" if has_atestado else "ABONO"
                                    ws.cell(row=row_idx, column=12, value=val_str)
                                    ws.cell(row=row_idx, column=13, value=val_str)
                                    ws.cell(row=row_idx, column=14, value=val_str)
                                    # For justified absences, balance impact is 0
                                    ws.cell(row=row_idx, column=16, value=0)
                                else:
                                    # For compensation, we show the label in main columns
                                    val_str = "COMPENSAÇÃO"
                                    ws.cell(row=row_idx, column=12, value=val_str)
                                    ws.cell(row=row_idx, column=13, value=val_str)
                                    ws.cell(row=row_idx, column=14, value=val_str)
                                    # BUT we subtract the correct deficit from the bank
                                    ws.cell(row=row_idx, column=16, value=-(deficit_sec / 86400.0)).number_format = '[h]:mm:ss'
                                
                        # After processing all days, link the balances between months
                        TEMPLATE_CELLS = {
                            'JAN': {'ant': 'M53', 'tot': 'M55'}, 'FEV': {'ant': 'M51', 'tot': 'M53'}, 
                            'MAR': {'ant': 'M53', 'tot': 'M55'}, 'ABR': {'ant': 'M52', 'tot': 'M54'}, 
                            'MAIO': {'ant': 'M53', 'tot': 'M55'}, 'JUN': {'ant': 'M52', 'tot': 'M54'}, 
                            'JUL': {'ant': 'M53', 'tot': 'M55'}, 'AGO': {'ant': 'M53', 'tot': 'M55'}, 
                            'SET': {'ant': 'M52', 'tot': 'M54'}, 'OUT': {'ant': 'M53', 'tot': 'M55'}, 
                            'NOV': {'ant': 'M52', 'tot': 'M54'}, 'DEZ': {'ant': 'M53', 'tot': 'M55'}
                        }
                        if m_name in TEMPLATE_CELLS:
                            ant_cell = TEMPLATE_CELLS[m_name]['ant']
                            
                            # Determine if this month is before or equal to the first punch month
                            first_punch_month = None
                            if user_records:
                                ts_first = rf(user_records[-1], 'timestamp') # user_records is reverse chronological, so [-1] is the first punch!
                                if ts_first:
                                    dt_first = ts_first if isinstance(ts_first, datetime.datetime) else datetime.datetime.strptime(str(ts_first).split('.')[0], '%Y-%m-%d %H:%M:%S')
                                    first_punch_month = datetime.date(dt_first.year, dt_first.month, 1)
                            
                            this_month_date = datetime.date(m_year, m_num, 1)
                            
                            is_before_or_start_month = False
                            if first_punch_month and this_month_date <= first_punch_month:
                                is_before_or_start_month = True
                            
                            if is_before_or_start_month:
                                ws['Q4'] = 0
                                ws[ant_cell] = f"=Q4"
                                ws['Q4'].number_format = '[h]:mm:ss'
                                ws[ant_cell].number_format = '[h]:mm:ss'
                            else:
                                if m_idx > 0:
                                    prev_name = month_names[m_idx - 1]
                                    prev_tot = TEMPLATE_CELLS[prev_name]['tot']
                                    ws['Q4'] = f"='{prev_name}'!{prev_tot}"
                                    ws[ant_cell] = f"=Q4"
                                else:
                                    prev_bal_days = get_previous_years_balance(user_records, m_year, daily_hours)
                                    ws['Q4'] = prev_bal_days
                                    ws[ant_cell] = f"=Q4"
                                    ws['Q4'].number_format = '[h]:mm:ss'
                                    ws[ant_cell].number_format = '[h]:mm:ss'
                                
            except Exception as e_main:
                import traceback, os
                with open(os.path.join(os.path.abspath(os.path.dirname(__file__)), "backend.log"), "a") as flog:
                    flog.write(f"\nCRITICAL EXCEL ERROR:\n{traceback.format_exc()}\n")
                raise e_main
        else:
            wb.remove(wb.active)
            groups = {}
            for r in rows:
                k = (rf(r,'matricula'), rf(r,'name'))
                groups.setdefault(k, []).append(r)
            for (m, n), items in groups.items():
                ws = wb.create_sheet(title=(n or m or "User")[:30])
                ws.append(["Matricula", "Nome", "Cargo", "Tipo", "Data/Hora", "Bairro", "Cidade", "Latitude", "Longitude", "Precisão (m)", "Endereço Completo", "Manual?", "Justificativa", "Anexo"])
                for r in items:
                    mat = rf(r,'matricula')
                    c = cargo_map.get(mat, 'Funcionario')
                    ws.append([mat, rf(r,'name'), c, rf(r,'record_type'), rf(r,'timestamp'), rf(r,'neighborhood'), rf(r,'city'), rf(r,'latitude'), rf(r,'longitude'), rf(r,'accuracy'), rf(r,'full_address'), 'Sim' if rf(r,'is_retroactive') else 'Não', rf(r,'justification') or '', rf(r,'document_path') or ''])
        
        if is_protected:
            # Fetch password from SystemConfig
            excel_pass = 'Sedu@2023'
            try:
                conn_tmp = get_db_connection()
                cur_tmp = conn_tmp.cursor()
                nolock = "" if isinstance(conn_tmp, sqlite3.Connection) else "WITH (NOLOCK)"
                cur_tmp.execute(f"SELECT value FROM SystemConfig {nolock} WHERE key = 'excel_protection_password'")
                row_tmp = cur_tmp.fetchone()
                if row_tmp:
                    excel_pass = rf(row_tmp, 'value')
                conn_tmp.close()
            except: pass
            
            from openpyxl.styles import Protection
            prot = Protection(locked=True)
            for sheet in wb.worksheets:
                for row in sheet.iter_rows():
                    for cell in row:
                        cell.protection = prot
                sheet.protection.sheet = True
                sheet.protection.password = excel_pass
        
        out = BytesIO()
        wb.save(out)
        out.seek(0)
        
        fname = "Relatorio_Geral.xlsx"
        if target_user_id and len(rows) > 0:
            user_n = rf(rows[0], 'name') or target_user_id
            fname = f"Ponto - {user_n}.xlsx"
            
        return send_file(out, download_name=fname, as_attachment=True)
    except Exception as e:
        import traceback
        traceback.print_exc()
        print(f"Excel Error: {e}")
        return jsonify({'message': str(e)}), 500
    finally:
        try: conn.close()
        except: pass

@app.route('/api/user/report', methods=['GET'])
@token_required
def get_user_self_report_excel(curr_user_mat, role):
    # Find user ID from matricula
    conn = get_db_connection()
    ph = get_ph(conn)
    cur = conn.cursor()
    cur.execute(f"SELECT id FROM Users WHERE matricula = {ph}", (curr_user_mat,))
    u = cur.fetchone()
    if not u:
        conn.close()
        return jsonify({'message': 'User not found'}), 404
    uid = rf(u, 'id')
    conn.close()
    
    year_arg = request.args.get('year')
    return _generate_excel_response(uid, year_arg, is_protected=True)

@app.route('/api/admin/export', methods=['GET'])
@token_required
def export_excel_legacy(curr_user_mat, role):
    return get_admin_report_excel(curr_user_mat, role)

@app.route('/api/admin/sync_all', methods=['POST'])
@token_required
def sync_all_users_admin(curr_user_mat, role):
    if role != 'admin':
        return jsonify({'message': 'Unauthorized'}), 401
    refresh_local_users()
    return jsonify({'message': 'Sincronização de usuários solicitada'}), 200

def refresh_local_users():
    """Helper to pull users from SQL into local SQLite."""
    try:
        conn = get_db_connection()
        if isinstance(conn, sqlite3.Connection):
            return # Already strictly local
        cursor = conn.cursor()
        cursor.execute("SELECT matricula, password, name, role, cargo FROM Users WITH (NOLOCK)")
        users = cursor.fetchall()
        
        sconn = sqlite3.connect(sqlite_path)
        scur = sconn.cursor()
        ensure_sqlite_schema(sconn)
        for u in users:
            scur.execute("INSERT OR REPLACE INTO Users (matricula, password, name, role, cargo) VALUES (?, ?, ?, ?, ?)", 
                         (rf(u,'matricula'), rf(u,'password'), rf(u,'name'), rf(u,'role'), rf(u,'cargo')))
        sconn.commit()
        sconn.close()
        conn.close()
    except: pass

@app.route('/api/sync', methods=['POST'])
@token_required
def sync_now(curr_user_mat, role):
    migrated, errs = perform_sync_for_user(curr_user_mat)
    return jsonify({'message': f'Sincronização concluída. {migrated} registros enviados.', 'migrated': migrated, 'errors': errs}), 200

def perform_sync_for_user(user_matricula):
    """
    Core sync logic that can be called via API or background thread.
    Returns (migrated_count, errors_list)
    """
    lock = get_user_sync_lock(user_matricula)
    if not lock.acquire(blocking=False):
        print(f"DEBUG: Sync already in progress for {user_matricula}. Skipping redundant run.")
        return 0, ["Sync already in progress"]

    try:
        migrated = 0
        errs = []
        
        # 1. Local-only move: OfflineQueue -> Local Mirror (TimeRecords)
        # This is always done first to empty the frontend-style queue into the local mirror.
        try:
            lconn = sqlite3.connect(sqlite_path)
            lconn.row_factory = sqlite3.Row
            ensure_sqlite_schema(lconn)
            lcur = lconn.cursor()
            
            # Fetch pending from Queue
            lcur.execute("SELECT * FROM OfflineQueue WHERE matricula = ? ORDER BY timestamp ASC", (user_matricula,))
            q_rows = lcur.fetchall()
            
            if q_rows:
                # Existing local signatures to avoid double local moves
                lcur.execute("SELECT record_type, timestamp, transaction_id FROM TimeRecords WHERE matricula = ?", (user_matricula,))
                local_sigs = set()
                for lr in lcur.fetchall():
                    ts_val = rf(lr, 'timestamp')
                    ts_str = str(ts_val).split('.')[0]
                    tx_id = rf(lr, 'transaction_id')
                    local_sigs.add((rf(lr, 'record_type'), ts_str))
                    if tx_id: local_sigs.add(str(tx_id))
                
                for qr in q_rows:
                    ts_val = rf(qr, 'timestamp')
                    ts_str = str(ts_val).split('.')[0]
                    tx_id = rf(qr, 'transaction_id')
                    tx_id_str = str(tx_id) if tx_id else None
                    
                    is_dupe = (rf(qr, 'record_type'), ts_str) in local_sigs or (tx_id_str and tx_id_str in local_sigs)
                    
                    # ATOMIC MOVE: Only delete from OfflineQueue if it's already in local mirror or successfully inserted now
                    can_delete = False
                    if is_dupe:
                        can_delete = True
                    else:
                        try:
                            # Read is_retroactive and justification safely using rf
                            q_is_retro = rf(qr, 'is_retroactive') or 0
                            q_justif = rf(qr, 'justification')
                            lcur.execute("""
                                INSERT INTO TimeRecords (user_id, matricula, user_name, record_type, timestamp, neighborhood, city, latitude, longitude, accuracy, full_address, transaction_id, is_retroactive, justification)
                                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                            """, (rf(qr,'user_id'), rf(qr,'matricula'), rf(qr,'user_name'), rf(qr,'record_type'), rf(qr,'timestamp'), 
                                  rf(qr,'neighborhood'), rf(qr,'city'), rf(qr,'latitude'), rf(qr,'longitude'), rf(qr,'accuracy'), rf(qr,'full_address'), tx_id, q_is_retro, q_justif))
                            can_delete = True
                        except Exception as ins_err:
                            errs.append(f"Local Insert failed for {tx_id}: {ins_err}")
                    
                    if can_delete:
                        lcur.execute("DELETE FROM OfflineQueue WHERE id = ?", (rf(qr, 'id'),))
                lconn.commit()
            lconn.close()
        except Exception as e:
            errs.append(f"Local move error: {e}")

        # 2. Remote Sync: Local Mirror -> SQL Server
        # Only if SQL Server is configured and reachable.
        remote_conn = get_remote_db_connection()
        if remote_conn:
            try:
                rcur = remote_conn.cursor()
                sph = get_ph(remote_conn)
                is_pymssql = hasattr(remote_conn, 'as_dict')
                
                # Ensure remote schema
                print(f"DEBUG: Checking remote schema for {user_matricula}...")
                try:
                    for col, col_type in [
                        ("latitude", "FLOAT"), ("longitude", "FLOAT"), ("accuracy", "FLOAT"), 
                        ("full_address", "NVARCHAR(MAX)"), ("transaction_id", "NVARCHAR(100)"),
                        ("matricula", "NVARCHAR(50)"), ("user_name", "NVARCHAR(200)"),
                        ("neighborhood", "NVARCHAR(200)"), ("city", "NVARCHAR(200)"),
                        ("is_retroactive", "INT"), ("justification", "NVARCHAR(MAX)")
                    ]:
                        try: 
                            rcur.execute(f"ALTER TABLE TimeRecords ADD {col} {col_type}")
                            if not is_pymssql and not getattr(remote_conn, 'autocommit', False): 
                                remote_conn.commit()
                        except: pass
                    print("DEBUG: Remote schema check complete.")
                except Exception as sch_err:
                    print(f"DEBUG: Schema check error: {sch_err}")

                # Fetch user info from Remote
                sql_user_id, user_name = get_user_info_by_matricula(user_matricula, remote_conn)
                print(f"DEBUG: Remote user fetch for {user_matricula}: ID={sql_user_id}, Name={user_name}")
                
                # v6.7: Sync user to remote if missing
                if not sql_user_id:
                    print(f"DEBUG: User {user_matricula} not found on remote. Attempting to sync user record...")
                    try:
                        ltmp_conn = sqlite3.connect(sqlite_path); ltmp_conn.row_factory = sqlite3.Row
                        ucur = ltmp_conn.cursor()
                        ucur.execute("SELECT * FROM Users WHERE matricula = ?", (str(user_matricula),))
                        u_data = ucur.fetchone()
                        if u_data:
                            # Use matricula as ID if numeric, else let it be auto (if possible)
                            try: fallback_id = int(str(user_matricula).strip())
                            except: fallback_id = None
                            
                            # SQL Server INSERT for Users
                            try:
                                # Ensure remote Users table has matricula if missing
                                try: rcur.execute("ALTER TABLE Users ADD matricula NVARCHAR(50)")
                                except: pass
                                if not is_pymssql and not getattr(remote_conn, 'autocommit', False): remote_conn.commit()

                                # Try inserting with explicit ID if it's numeric and we have identity insert maybe? 
                                # Simpler: just insert and let remote handle ID, we use matricula for mapping mostly.
                                rcur.execute(f"INSERT INTO Users (name, matricula, password, role) VALUES ({sph}, {sph}, {sph}, {sph})", 
                                             (rf(u_data, 'name'), str(user_matricula), rf(u_data, 'password'), rf(u_data, 'role')))
                                if not is_pymssql and not getattr(remote_conn, 'autocommit', False): remote_conn.commit()
                                print(f"DEBUG: Successfully synced user {user_matricula} to remote.")
                                # Re-fetch ID
                                sql_user_id, user_name = get_user_info_by_matricula(user_matricula, remote_conn)
                            except Exception as u_ins_err:
                                print(f"DEBUG: Could not insert user to remote: {u_ins_err}")
                        ltmp_conn.close()
                    except Exception as u_sync_err:
                        print(f"DEBUG: User sync logic failed: {u_sync_err}")

                # v6.6: Improved Fallback - use matricula as integer ID if remote fetch still fails
                if not user_name:
                    try:
                        ltmp_conn = sqlite3.connect(sqlite_path)
                        ltmp_conn.row_factory = sqlite3.Row
                        _, local_user_name = get_user_info_by_matricula(user_matricula, ltmp_conn)
                        ltmp_conn.close()
                        if local_user_name:
                            user_name = local_user_name
                            print(f"DEBUG: Using local info fallback for {user_matricula}: {user_name}")
                        
                        if not sql_user_id:
                            try:
                                # Many systems use matricula as the primary/external ID
                                sql_user_id = int(str(user_matricula).strip())
                                print(f"DEBUG: Using matricula as numeric fallback for user_id: {sql_user_id}")
                            except:
                                sql_user_id = 0
                                print("DEBUG: Using 0 as ultimate user_id fallback.")
                    except Exception as f_err:
                        print(f"DEBUG: Fallback error: {f_err}")
                
                # Build remote signatures
                existing_sigs = set()
                # Use a small check range for speed, but deep enough to catch recent dupes (last 90 days)
                # SQL Server specific syntax for DATEADD if not sqlite
                nolock = "WITH (NOLOCK)" if not is_pymssql else ""
                rcur.execute(f"SELECT record_type, timestamp, transaction_id FROM TimeRecords {nolock} WHERE matricula = {sph} AND timestamp >= DATEADD(day, -90, GETDATE())", (user_matricula,))
                for row in rcur.fetchall():
                    ts_val = rf(row, 'timestamp')
                    ts_str = str(ts_val).split('.')[0]
                    tx_id = rf(row, 'transaction_id')
                    existing_sigs.add((rf(row, 'record_type'), ts_str))
                    if tx_id: existing_sigs.add(str(tx_id))
                
                # Fetch from local mirror
                lconn = sqlite3.connect(sqlite_path)
                lconn.row_factory = sqlite3.Row
                lcur = lconn.cursor()
                lcur.execute("SELECT * FROM TimeRecords WHERE matricula = ? AND timestamp >= date('now', '-90 days')", (str(user_matricula),))
                mirror_rows = lcur.fetchall()
                print(f"DEBUG: Local mirror rows found for {user_matricula}: {len(mirror_rows)}")
                
                for mr in mirror_rows:
                    ts_val = rf(mr, 'timestamp')
                    # Standardize
                    ts_dt = ts_val
                    if isinstance(ts_val, str):
                        try:
                            if '.' in ts_val: ts_dt = datetime.datetime.strptime(ts_val, "%Y-%m-%d %H:%M:%S.%f")
                            else: ts_dt = datetime.datetime.strptime(ts_val, "%Y-%m-%d %H:%M:%S")
                        except: pass
                    if isinstance(ts_dt, datetime.datetime): ts_dt = ts_dt.replace(microsecond=0)
                    cmp_ts = str(ts_dt).split('.')[0]
                    tx_id = rf(mr, 'transaction_id')
                    tx_id_str = str(tx_id) if tx_id else None
                    
                    if (rf(mr, 'record_type'), cmp_ts) not in existing_sigs and (not tx_id_str or tx_id_str not in existing_sigs):
                        try:
                            m_is_retro = rf(mr, 'is_retroactive') or 0
                            m_justif = rf(mr, 'justification')
                            rcur.execute(f"""
                                INSERT INTO TimeRecords (user_id, matricula, user_name, record_type, timestamp, neighborhood, city, latitude, longitude, accuracy, full_address, transaction_id, is_retroactive, justification)
                                VALUES ({sph}, {sph}, {sph}, {sph}, {sph}, {sph}, {sph}, {sph}, {sph}, {sph}, {sph}, {sph}, {sph}, {sph})
                            """, (sql_user_id, str(user_matricula), user_name, rf(mr, 'record_type'), ts_dt,
                                  rf(mr, 'neighborhood'), rf(mr, 'city'), rf(mr, 'latitude'), rf(mr, 'longitude'), 
                                  rf(mr, 'accuracy'), rf(mr, 'full_address'), tx_id, m_is_retro, m_justif))
                            migrated += 1
                            print(f"DEBUG: Successfully synced record {tx_id} for {user_matricula}")
                        except Exception as e:
                            print(f"DEBUG: SQL Insert error for {tx_id}: {e}")
                            errs.append(f"SQL Insert error: {e}")
                
                if not is_pymssql and not getattr(remote_conn, 'autocommit', False): 
                    remote_conn.commit()
                remote_conn.close()
                lconn.close()
            except Exception as e:
                errs.append(f"Remote sync error: {e}")
        
        return migrated, errs
    finally:
        lock.release()

def auto_sync_all():
    """Finds all users with pending items and syncs them."""
    print("DEBUG: Starting automatic background synchronization...")
    try:
        sconn = sqlite3.connect(sqlite_path)
        sconn.row_factory = sqlite3.Row
        scur = sconn.cursor()
        # Find all distinct matriculas and user_ids in the queue
        scur.execute("SELECT DISTINCT matricula, user_id FROM OfflineQueue")
        rows = scur.fetchall()
        
        synced_users = set()
        total_migrated = 0
        
        for r in rows:
            m = rf(r, 'matricula')
            uid = rf(r, 'user_id')
            
            # If matricula is missing, try to find it via user_id
            if not m or m == '':
                scur.execute("SELECT matricula FROM Users WHERE id = ? OR matricula = ?", (uid, str(uid)))
                u_row = scur.fetchone()
                m = rf(u_row, 'matricula') if u_row else None
            
            if m and m not in synced_users:
                print(f"DEBUG: Auto-syncing for matricula: {m}")
                migrated, errs = perform_sync_for_user(m)
                total_migrated += migrated
                synced_users.add(m)
        
        sconn.close()
        if total_migrated > 0:
            print(f"DEBUG: Automatic sync complete. {total_migrated} records synchronized.")
    except Exception as e:
        print(f"DEBUG: Auto-sync error: {e}")

@app.after_request
def add_header(response):
    if 'text/html' in response.content_type:
        response.headers['Cache-Control'] = 'no-cache, no-store, must-revalidate'
        response.headers['Pragma'] = 'no-cache'
        response.headers['Expires'] = '0'
    return response

if __name__ == '__main__':
    port = int(os.getenv('PORT', '5005'))
    ensure_default_admin()
    migrate_local_data()
    start_health_check()
    app.run(host='0.0.0.0', debug=False, port=port)
